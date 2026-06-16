# Модуль загрузки данных

import openpyxl
import openpyxl.utils
import json

from django.db import transaction
from django.utils import timezone

from .common import *


# ============================== ENDPOINTS ============================== #

@method(POST)
@jsonResponse
@csrf_exempt
def import_excel(request):
    """
    Import data from an Excel file.
    Supported sheets:
    1. Связь ФИО и ID - маппинг RSV ID → ФИО
    2. Сравнение по компетенциям - результаты РСВ
    3. Мотивационный профиль
    4. Ценностный профиль
    5. Образовательные курсы
    6. Итоги успеваемости участников - данные по дисциплинам
    """
    excel_file = request.FILES.get("file")
    config_json = request.POST.get("config_json")

    if not excel_file:
        raise ResponseError("No file uploaded")
    if not config_json:
        raise ResponseError("Missing config_json")

    try:
        config = json.loads(config_json)
    except json.JSONDecodeError:
        raise ResponseError("Invalid JSON in config_json")

    wb = openpyxl.load_workbook(excel_file, data_only=True)

    created_count = 0
    updated_count = 0
    mapping_count = 0
    participants_created = 0
    participants_updated = 0

    with transaction.atomic():
        for sheet_info in config["sheets"]:
            sheet_name = sheet_info["name"]
            start_row = sheet_info["start_row"]
            columns = sheet_info["columns"]

            if sheet_name not in wb.sheetnames:
                debugPrint(f"[xls load] (!):  лист '{sheet_name}' отсутствует в файле")
                continue

            ws = wb[sheet_name]

            debugPrint(f"[xls load] (i): *** ОБРАБОТКА ЛИСТА '{sheet_name}' ***")

            # Ключевое поле для определения конца данных
            KEY_FIELD = {
                "Связь ФИО и ID":                "rsv_id",
                "Сравнение по компетенциям":     "part_rsv",
                "Мотивационный профиль":         "part_rsv",
                "Ценностный профиль":            "part_rsv",
                "Образовательные курсы":         "part_rsv",
                "Итоги успеваемости участников": "student_name",
            }
            key_field = KEY_FIELD.get(sheet_name)
            EMPTY_ROWS_LIMIT = 5  # остановиться после N пустых строк подряд

            row_count = 0
            empty_streak = 0
            for row in ws.iter_rows(min_row=start_row):

                def get_col(col_letter):
                    """Получает значение ячейки по букве колонки."""
                    if not col_letter:
                        return None
                    try:
                        col_idx = openpyxl.utils.column_index_from_string(col_letter) - 1
                        if col_idx >= len(row):
                            return None
                        return row[col_idx].value
                    except:
                        return None

                row_data = {k: get_col(v) for k, v in columns.items()}

                # Остановка по пустым строкам: если ключевое поле пусто N раз подряд — конец данных
                if key_field:
                    key_value = clean_value(row_data.get(key_field))
                    if not key_value:
                        empty_streak += 1
                        if empty_streak >= EMPTY_ROWS_LIMIT:
                            debugPrint(f"[xls load] (i): {EMPTY_ROWS_LIMIT} пустых строк подряд — завершение; {row_count} строк обработано")
                            break
                        continue
                    else:
                        empty_streak = 0

                # ============================================================
                # ЛИСТ 1: "Связь ФИО и ID"
                # ============================================================
                if sheet_name == "Связь ФИО и ID":
                    rsv_id = clean_value(row_data.get("rsv_id"))
                    student_name = clean_value(row_data.get("student_name"))
                    student_gender = parse_gender(row_data.get("student_gender"))  # ИЗМЕНЕНО: теперь INT
                    email = clean_value(row_data.get("email"))

                    if not rsv_id or not student_name:
                        continue

                    # ИЗМЕНЕНО: новые имена полей
                    mapping, created = StudentMapping.objects.update_or_create(
                        mapping_rsv=str(rsv_id),  # было rsv_id
                        defaults={
                            'mapping_stud_name': student_name,    # было student_name
                            'mapping_stud_gender': student_gender, # было student_gender, теперь INT
                            'mapping_email': email,                # было email
                            'mapping_created_at': timezone.now()   # добавлено
                        }
                    )

                    mapping_count += 1
                    row_count += 1

                    if row_count % 100 == 0:
                        debugPrint(f"[xls load] (i): обработано {row_count} связей")

                    continue

                # ============================================================
                # ЛИСТ 2: "Сравнение по компетенциям" (РСВ)
                # ============================================================
                if sheet_name == "Сравнение по компетенциям":
                    center_name = clean_value(row_data.get("center_name"))
                    if not center_name:
                        continue

                    center, _ = CompetenceCenters.objects.get_or_create(center_name=center_name)

                    institution = None
                    if (inst_name := clean_value(row_data.get("inst_name"))):
                        institution, _ = Institutions.objects.get_or_create(inst_name=inst_name)

                    edu_level = None
                    if (edu_level_name := clean_value(row_data.get("edu_level_name"))):
                        edu_level, _ = EducationLevels.objects.get_or_create(edu_level_name=edu_level_name)

                    # ИЗМЕНЕНО: form → edu_form
                    edu_form = None
                    if (form_name := clean_value(row_data.get("form_name"))):
                        edu_form, _ = EducationForms.objects.get_or_create(edu_form_name=form_name)

                    # ИЗМЕНЕНО: spec → edu_specialty
                    edu_specialty = None
                    if (spec_name := clean_value(row_data.get("spec_name"))):
                        edu_specialty, _ = EducationSpecialties.objects.get_or_create(edu_spec_name=spec_name)

                    # part_rsv - это ID из тестирования РСВ
                    rsv = clean_value(row_data.get("part_rsv"))  # ИЗМЕНЕНО: part_rsv вместо part_rsv_id
                    if not rsv:
                        continue

                    # Получаем или создаём участника, сразу заполняя все part_* поля
                    # ИЗМЕНЕНО: Participants теперь содержит ТОЛЬКО базовую информацию
                    participant, created = Participants.objects.get_or_create(
                        part_rsv=str(rsv),  # ИЗМЕНЕНО: part_rsv вместо part_rsv_id
                        defaults={
                            'part_gender': parse_gender(row_data.get("part_gender")),
                            'part_course_num': clean_value(row_data.get("part_course_num"), int),
                        }
                    )

                    if created:
                        participants_created += 1
                    else:
                        # Обновляем только если изменилось (без учебной информации!)
                        Participants.objects.filter(pk=participant.pk).update(
                            part_gender=parse_gender(row_data.get("part_gender")),
                            part_course_num=clean_value(row_data.get("part_course_num"), int),
                        )
                        participants_updated += 1

                    # Создаём или обновляем результат
                    # ИЗМЕНЕНО: TestResults теперь содержит ВСЮ учебную информацию
                    year = clean_value(row_data.get("res_year"))
                    # ВНИМАНИЕ: unique_together теперь (res_participant, res_year) без course_num
                    # course_num теперь отдельное поле в TestResults, но не входит в уникальность

                    # Пытаемся найти существующий результат
                    result, created_res = TestResults.objects.update_or_create(
                        res_participant=participant,
                        res_year=year,
                        defaults={
                            tRES.CENTER: center,
                            tRES.INSTITUTION: institution,
                            tRES.EDU_LEVEL: edu_level,
                            tRES.EDU_FORM: edu_form,           # ИЗМЕНЕНО: edu_form
                            tRES.EDU_SPEC: edu_specialty,      # ИЗМЕНЕНО: edu_specialty
                            tRES.COURSE_NUM: clean_value(row_data.get("res_course_num"), int),  # ИЗМЕНЕНО: course_num
                            tRES.POTENTIAL: clean_value(row_data.get("res_high_potential"), int),  # ИЗМЕНЕНО: potential
                            tRES.REPORT: clean_value(row_data.get("res_summary_report")),       # ИЗМЕНЕНО: report
                        }
                    )

                    # Обновляем компетенции (только если есть значения)
                    comp_updates = {}
                    for comp in COMP.list:
                        value = clean_value(row_data.get(comp), int)
                        if value is not None:
                            comp_updates[comp] = value
                    
                    if comp_updates:
                        TestResults.objects.filter(pk=result.pk).update(**comp_updates)

                    created_count += 1 if created_res else 0
                    updated_count += 0 if created_res else 1
                    row_count += 1

                    if row_count % 100 == 0:
                        debugPrint(f"[xls load] (i): обработано {row_count} результатов РСВ")

                # ============================================================
                # ЛИСТ 3: "Мотивационный профиль"
                # ============================================================
                elif sheet_name == "Мотивационный профиль":
                    rsv = clean_value(row_data.get("part_rsv"))  # ИЗМЕНЕНО: part_rsv
                    if not rsv:
                        continue

                    try:
                        participant = Participants.objects.get(part_rsv=str(rsv))  # ИЗМЕНЕНО: part_rsv
                    except Participants.DoesNotExist:
                        debugPrint(f"[xls load] (!): участник RSV {rsv} не найден")
                        continue

                    year = clean_value(row_data.get("res_year"))

                    # Обновляем мотиваторы
                    mot_updates = {}
                    for mot in MOT.list:
                        value = clean_value(row_data.get(mot), float)
                        if value is not None:
                            mot_updates[mot] = value
                    
                    if mot_updates:
                        updated_rows = TestResults.objects.filter(
                            res_participant=participant, 
                            res_year=year
                        ).update(**mot_updates)
                        updated_count += updated_rows
                    
                    row_count += 1

                # ============================================================
                # ЛИСТ 4: "Ценностный профиль"
                # ============================================================
                elif sheet_name == "Ценностный профиль":
                    rsv = clean_value(row_data.get("part_rsv"))  # ИЗМЕНЕНО: part_rsv
                    if not rsv:
                        continue

                    try:
                        participant = Participants.objects.get(part_rsv=str(rsv))
                    except Participants.DoesNotExist:
                        debugPrint(f"[xls load] (!): участник RSV {rsv} не найден")
                        continue

                    year = clean_value(row_data.get("res_year"))

                    # Обновляем ценности
                    val_updates = {}
                    for val in VAL.list:
                        value = clean_value(row_data.get(val), int)
                        if value is not None:
                            val_updates[val] = value
                    
                    if val_updates:
                        updated_rows = TestResults.objects.filter(
                            res_participant=participant, 
                            res_year=year
                        ).update(**val_updates)
                        updated_count += updated_rows
                    
                    row_count += 1

                # ============================================================
                # ЛИСТ 5: "Образовательные курсы"
                # ============================================================
                elif sheet_name == "Образовательные курсы":
                    rsv = clean_value(row_data.get("part_rsv"))  # ИЗМЕНЕНО: part_rsv
                    if not rsv:
                        continue

                    try:
                        participant = Participants.objects.get(part_rsv=str(rsv))
                    except Participants.DoesNotExist:
                        debugPrint(f"[xls load] (!): участник RSV {rsv} не найден")
                        continue

                    # ИЗМЕНЕНО: курс теперь CourseResults, не связан с year
                    # Используем update_or_create
                    course_defaults = {}
                    for cur in CUR.list:
                        value = clean_value(row_data.get(cur), float)
                        if value is not None:
                            course_defaults[cur] = value
                    
                    if course_defaults:
                        course_result, created_res = CourseResults.objects.update_or_create(
                            course_participant=participant,  # ИЗМЕНЕНО: participant, а не participant.part_id
                            defaults=course_defaults
                        )
                        if created_res:
                            created_count += 1
                        else:
                            updated_count += 1
                    
                    row_count += 1

                # ============================================================
                # ЛИСТ 6: "Итоги успеваемости участников" (ДИСЦИПЛИНЫ)
                # ============================================================
                elif sheet_name == "Итоги успеваемости участников":
                    student_name = clean_value(row_data.get("student_name"))
                    if not student_name:
                        continue

                    # Ищем RSV ID по ФИО в таблице маппинга
                    try:
                        mapping = StudentMapping.objects.get(mapping_stud_name=student_name)  # ИЗМЕНЕНО
                        rsv = mapping.mapping_rsv  # ИЗМЕНЕНО
                    except StudentMapping.DoesNotExist:
                        debugPrint(f"[xls load] (!): ФИО '{student_name}' не найдено в маппинге")
                        continue
                    except StudentMapping.MultipleObjectsReturned:
                        debugPrint(f"[xls load] (!): несколько записей для ФИО '{student_name}'")
                        mapping = StudentMapping.objects.filter(mapping_stud_name=student_name).first()
                        rsv = mapping.mapping_rsv

                    # Ищем участника по RSV
                    try:
                        participant = Participants.objects.get(part_rsv=rsv)  # ИЗМЕНЕНО: part_rsv
                    except Participants.DoesNotExist:
                        debugPrint(f"[xls load] (!): участник RSV {rsv} не найден")
                        continue

                    year = clean_value(row_data.get("perf_year"))
                    discipline_name = clean_value(row_data.get("perf_discipline"))

                    if not year or not discipline_name:
                        continue

                    # ИЗМЕНЕНО: дисциплины теперь в отдельной таблице
                    discipline, _ = EducationDisciplines.objects.get_or_create(
                        edu_disc_name=discipline_name
                    )

                    # ИЗМЕНЕНО: обновленная структура AcademicPerformances
                    # Оценки теперь числа (1-5)
                    main_grade = convert_grade_to_int(clean_value(row_data.get("perf_main_attestation")))
                    
                    perf_defaults = {
                        'perf_current': clean_value(row_data.get("perf_current"), float),
                        'perf_digital': clean_value(row_data.get("perf_digital"), float),
                        'perf_main': main_grade,
                        # Добавляем новые поля, если есть в Excel
                        'perf_first_retake': convert_grade_to_int(clean_value(row_data.get("perf_first_retake"))),
                        'perf_second_retake': convert_grade_to_int(clean_value(row_data.get("perf_second_retake"))),
                        'perf_grade_retake': convert_grade_to_int(clean_value(row_data.get("perf_grade_retake"))),
                        'perf_final': convert_grade_to_int(clean_value(row_data.get("perf_final"))),
                    }
                    
                    # Удаляем None значения
                    perf_defaults = {k: v for k, v in perf_defaults.items() if v is not None}

                    performance, created_perf = AcademicPerformances.objects.update_or_create(
                        perf_participant=participant,
                        perf_edu_discipline=discipline,
                        perf_year=year,
                        defaults=perf_defaults
                    )

                    if created_perf:
                        created_count += 1
                    else:
                        updated_count += 1
                    
                    row_count += 1
                    
                    if row_count % 100 == 0:
                        debugPrint(f"[xls load] (i): обработано {row_count} записей дисциплин")

            debugPrint(f"[xls load] (i): лист '{sheet_name}' завершён; {row_count} строк")

    debugPrint(f"[xls load] (i): импорт завершён")
    debugPrint(f"[xls load] (i): создано записей: {created_count}")
    debugPrint(f"[xls load] (i): обновлено записей: {updated_count}")
    debugPrint(f"[xls load] (i): связей ФИО→ID: {mapping_count}")
    debugPrint(f"[xls load] (i): создано участников: {participants_created}")
    debugPrint(f"[xls load] (i): обновлено участников: {participants_updated}")

    return {
        "status": "success", 
        "created": created_count, 
        "updated": updated_count,
        "mapped": mapping_count,
        "participants_created": participants_created,
        "participants_updated": participants_updated
    }


# fixme: get rid of this endpoint, it should be hardcoded in frontend
@method(GET)
@jsonResponse
@csrf_exempt
def get_expected_fields(request):
    """ Return expected field for each sheet.
    """
    return {
        "Связь ФИО и ID": [
            "rsv_id",           # mapping_rsv
            "student_name",     # mapping_stud_name
            "student_gender",   # mapping_stud_gender (конвертится в INT)
            "email"             # mapping_email
        ],
        "Сравнение по компетенциям": [
            "center_name", "inst_name", "edu_level_name", "form_name", "spec_name",
            "part_rsv", "part_gender", "part_course_num",
            "res_year", "res_course_num", "res_high_potential", "res_summary_report",
            *COMP.list
        ],
        "Мотивационный профиль": ["part_rsv", "res_year", *MOT.list],
        "Ценностный профиль": ["part_rsv", "res_year", *VAL.list],
        "Образовательные курсы": ["part_rsv", *CUR.list],
        "Итоги успеваемости участников": [
            "student_name", "perf_year", "perf_discipline",
            "perf_current", "perf_digital", "perf_main_attestation",
            "perf_first_retake", "perf_second_retake", "perf_grade_retake", "perf_final"
        ]
    }


@method(GET)
@jsonResponse
@csrf_exempt
def get_templates(request):
    """ Get list of all data load templates.
    """
    templates = DataUploadTemplate.objects.all().values(
        'template_id',      # было 'id'
        'template_name',    # было 'name'
        'template_description',  # было 'description'
        'template_config',       # было 'config'
        'template_updated_at'    # было 'updated_at'
    )
    
    # Переименовываем поля для обратной совместимости с фронтендом
    renamed_templates = []
    for t in templates:
        renamed_templates.append({
            'id': t['template_id'],
            'name': t['template_name'],
            'description': t['template_description'],
            'config': t['template_config'],
            'updated_at': t['template_updated_at']
        })
    
    return {'templates': renamed_templates}


@method(POST)
@jsonResponse
@csrf_exempt
def save_template(request):
    """ Save or update data load template.
    """
    data = json.loads(request.body)
    name = data.get('name')
    config = data.get('config')
    description = data.get('description', '')

    if not name:
        raise ResponseError("Missing template name")
    if not config:
        raise ResponseError("Missing config")

    # Используем новые имена полей
    template, created = DataUploadTemplate.objects.update_or_create(
        template_name=name,  # было 'name'
        defaults={
            'template_config': config,           # было 'config'
            'template_description': description, # было 'description'
            'template_updated_at': timezone.now()
        }
    )

    # Если только создали, устанавливаем created_at
    if created:
        template.template_created_at = timezone.now()
        template.save()

    return {
        'id': template.template_id,        # было template.id
        'name': template.template_name,    # было template.name
        'created': created
    }


@method(DELETE)
@jsonResponse
@csrf_exempt
def delete_template(request, template_id):
    """ Delete data load template.
    """
    try:
        # Используем template_id как первичный ключ
        template = DataUploadTemplate.objects.get(template_id=template_id)
        template.delete()
        return {'status': 'deleted', 'id': template_id}
    except DataUploadTemplate.DoesNotExist:
        raise ResponseError(f"Template with id {template_id} not found", status=404)


# ============================== UTILITIES ============================== #

def parse_gender(value) -> int | None:
    """ Конвертирует обозначение пола в число: М/Муж/Male → 1, Ж/Жен/Female → 2.
    """
    if value is None:
        return None
    s = str(value).strip().lower()
    if s in ("м", "муж", "мужской", "male", "m", "1"):
        return 1
    if s in ("ж", "жен", "женский", "female", "f", "2"):
        return 2
    return None


def clean_value(value, value_type: type[str] | type[int] | type[float] = str) -> str | int | float | None:
    """ Преобразует значение из Excel ячейки в нужный формат.
    """
    if value is None:
        return None

    # Если это уже нужный тип
    if isinstance(value, value_type) and value_type != str:
        return value

    s = str(value).strip()

    # Маркеры пустых значений
    if s.lower() in ("", "-", "–", "—", "отсутствует", "н/д", "н/п", "нет", "na", "n/a", "null", "none"):
        return None
    
    try:
        if value_type == int:
            # Обработка float в int
            if isinstance(value, float):
                return int(value) if not pd.isna(value) else None
            return int(float(s.replace(",", ".")))
        
        if value_type == float:
            if isinstance(value, (int, float)):
                return float(value)
            return float(s.replace(",", "."))
    
    except (ValueError, TypeError):
        return None

    return s or None


def convert_grade_to_int(grade_str: str | None) -> int | None:
    """ Конвертирует текстовую оценку в число (1-5).
        'отл.' → 5
        'хор.' → 4
        'удовл.' → 3
        'неудовл.' → 2
        'не явился' → 1
    """
    if grade_str is None:
        return None
    
    s = str(grade_str).strip().lower()
    
    grade_map = {
        'отл.': 5,
        'отлично': 5,
        'хор.': 4,
        'хорошо': 4,
        'удовл.': 3,
        'удовлетворительно': 3,
        'неудовл.': 2,
        'неудовлетворительно': 2,
        'не явился': 1,
        'неявка': 1,
    }
    
    # Пробуем прямой маппинг
    if s in grade_map:
        return grade_map[s]
    
    # Пробуем как число
    try:
        num = int(float(s))
        if 1 <= num <= 5:
            return num
    except:
        pass
    
    return None
