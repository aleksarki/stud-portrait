import json
import openpyxl

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count, Avg, Min
from django.db.models import Q
from django.db import transaction
from django.utils import timezone
import uuid
import pandas as pd
import numpy as np
from collections import defaultdict

from .models import (
    Participants, Competencecenters as CompetenceCenters, Institutions,
    Educationlevels, Studyforms, Specialties,
    Results, Course, Academicperformance
)

# Create your views here.

successResponse = lambda d: JsonResponse({**{"status": "success"}, **d})
errorResponse = lambda m: JsonResponse({"status": "error", "message": m}, status=400)
notFoundResponse = lambda m: JsonResponse({"status": "error", "message": m}, status=404)
notAllowedResponse = lambda: JsonResponse({"status": "error", "message": "Method not allowed"}, status=405)
exceptionResponse = lambda e: JsonResponse({"status": "error", "message": str(e)}, status=500)


def method(method):
    def inner(func):
        @csrf_exempt
        def wrapper(request):
            if request.method != method:
                return notAllowedResponse()
            return func(request)
        return wrapper
    return inner


def response(func):
    @csrf_exempt
    def wrapper(request):
        try:
            return successResponse(func(request))
        except Exception as e:
            return exceptionResponse(e)
    return wrapper


def clean_value(value, value_type="str"):
    """
    Преобразует данные из Excel к нормальному виду:
    - заменяет "-", "Отсутствует" и т.п. на None
    - преобразует числа из строк (в т.ч. с запятой)
    - убирает пробелы
    """
    if value is None:
        return None

    # Приводим к строке и чистим пробелы
    s = str(value).strip()

    # Маркеры отсутствующих значений
    empty_markers = {"", "-", "–", "—", "Отсутствует", "н/д", "н/п", "нет", "na", "n/a"}
    if s.lower() in {m.lower() for m in empty_markers}:
        return None

    # Тип int
    if value_type == "int":
        try:
            return int(float(s.replace(",", ".")))
        except ValueError:
            return None

    # Тип float / decimal
    if value_type == "float":
        try:
            return float(s.replace(",", "."))
        except ValueError:
            return None

    # Строка
    return s or None


@method('GET')
@csrf_exempt
def results(request):
    part_id = request.GET.get('part_id')
    
    if not part_id:
        # Убираем пагинацию и загружаем все результаты
        results_query = Results.objects.all().select_related(
            'res_participant', 'res_center', 'res_institution',
            'res_edu_level', 'res_form', 'res_spec'
        ).exclude(
            res_course_num__isnull=True
        ).order_by(
            "res_participant_id",
            "res_course_num"
        )
        
        # Фильтры (оставляем для поиска)
        institution_filter = request.GET.get('institution')
        year_filter = request.GET.get('year')
        center_filter = request.GET.get('center')
        participant_filter = request.GET.get('participant')
        specialty_filter = request.GET.get('specialty')
        
        if institution_filter:
            results_query = results_query.filter(res_institution__inst_name__icontains=institution_filter)
        if year_filter:
            results_query = results_query.filter(res_year=year_filter)
        if center_filter:
            results_query = results_query.filter(res_center__center_name__icontains=center_filter)
        if participant_filter:
            results_query = results_query.filter(res_participant__part_name__icontains=participant_filter)
        if specialty_filter:
            results_query = results_query.filter(res_spec__spec_name__icontains=specialty_filter)
        
        # Убираем пагинацию - загружаем все результаты
        results_list = results_query
        
        results_data = []
        for result in results_list:
            result_data = {
                'res_id': result.res_id,
                'participant': {
                    'part_id': result.res_participant.part_id,
                    'part_name': result.res_participant.part_name,
                    'part_gender': result.res_participant.part_gender,
                },
                'center': result.res_center.center_name if result.res_center else None,
                'institution': result.res_institution.inst_name if result.res_institution else None,
                'edu_level': result.res_edu_level.edu_level_name if result.res_edu_level else None,
                'study_form': result.res_form.form_name if result.res_form else None,
                'specialty': result.res_spec.spec_name if result.res_spec else None,
                'res_year': result.res_year,
                'res_course_num': result.res_course_num,
                'res_high_potential': result.res_high_potential,
                'res_summary_report': result.res_summary_report,
                
                # Компетенции
                'competences': {
                    'res_comp_info_analysis': result.res_comp_info_analysis,
                    'res_comp_planning': result.res_comp_planning,
                    'res_comp_result_orientation': result.res_comp_result_orientation,
                    'res_comp_stress_resistance': result.res_comp_stress_resistance,
                    'res_comp_partnership': result.res_comp_partnership,
                    'res_comp_rules_compliance': result.res_comp_rules_compliance,
                    'res_comp_self_development': result.res_comp_self_development,
                    'res_comp_leadership': result.res_comp_leadership,
                    'res_comp_emotional_intel': result.res_comp_emotional_intel,
                    'res_comp_client_focus': result.res_comp_client_focus,
                    'res_comp_communication': result.res_comp_communication,
                    'res_comp_passive_vocab': result.res_comp_passive_vocab,
                },
                
                # Мотиваторы
                'motivators': {
                    'res_mot_autonomy': result.res_mot_autonomy,
                    'res_mot_altruism': result.res_mot_altruism,
                    'res_mot_challenge': result.res_mot_challenge,
                    'res_mot_salary': result.res_mot_salary,
                    'res_mot_career': result.res_mot_career,
                    'res_mot_creativity': result.res_mot_creativity,
                    'res_mot_relationships': result.res_mot_relationships,
                    'res_mot_recognition': result.res_mot_recognition,
                    'res_mot_affiliation': result.res_mot_affiliation,
                    'res_mot_self_development': result.res_mot_self_development,
                    'res_mot_purpose': result.res_mot_purpose,
                    'res_mot_cooperation': result.res_mot_cooperation,
                    'res_mot_stability': result.res_mot_stability,
                    'res_mot_tradition': result.res_mot_tradition,
                    'res_mot_management': result.res_mot_management,
                    'res_mot_work_conditions': result.res_mot_work_conditions,
                },
                
                # Ценности
                'values': {
                    'res_val_honesty_justice': result.res_val_honesty_justice,
                    'res_val_humanism': result.res_val_humanism,
                    'res_val_patriotism': result.res_val_patriotism,
                    'res_val_family': result.res_val_family,
                    'res_val_health': result.res_val_health,
                    'res_val_environment': result.res_val_environment,
                }
            }
            results_data.append(result_data)
            
        return successResponse({
            "results": results_data,
            "total_count": len(results_data)
        })
    
    try:
        part_id = int(part_id)
        
        try:
            participant = Participants.objects.select_related(
                'part_institution', 'part_spec', 'part_edu_level', 'part_form'
            ).get(part_id=part_id)
        except Participants.DoesNotExist:
            return notFoundResponse("Participant not found")
        
        results_list = []
        for result in Results.objects.filter(res_participant=part_id).select_related(
            'res_center', 'res_institution', 'res_edu_level', 'res_form', 'res_spec'
        ):
            result_data = {
                'res_id': result.res_id,
                'center': result.res_center.center_name if result.res_center else None,
                'institution': result.res_institution.inst_name if result.res_institution else None,
                'edu_level': result.res_edu_level.edu_level_name if result.res_edu_level else None,
                'study_form': result.res_form.form_name if result.res_form else None,
                'specialty': result.res_spec.spec_name if result.res_spec else None,
                'res_year': result.res_year,
                'res_course_num': result.res_course_num,
                'res_high_potential': result.res_high_potential,
                'res_summary_report': result.res_summary_report,
                
                # Компетенции
                'competences': {
                    'res_comp_info_analysis': result.res_comp_info_analysis,
                    'res_comp_planning': result.res_comp_planning,
                    'res_comp_result_orientation': result.res_comp_result_orientation,
                    'res_comp_stress_resistance': result.res_comp_stress_resistance,
                    'res_comp_partnership': result.res_comp_partnership,
                    'res_comp_rules_compliance': result.res_comp_rules_compliance,
                    'res_comp_self_development': result.res_comp_self_development,
                    'res_comp_leadership': result.res_comp_leadership,
                    'res_comp_emotional_intel': result.res_comp_emotional_intel,
                    'res_comp_client_focus': result.res_comp_client_focus,
                    'res_comp_communication': result.res_comp_communication,
                    'res_comp_passive_vocab': result.res_comp_passive_vocab,
                },
                
                # Мотиваторы
                'motivators': {
                    'res_mot_autonomy': result.res_mot_autonomy,
                    'res_mot_altruism': result.res_mot_altruism,
                    'res_mot_challenge': result.res_mot_challenge,
                    'res_mot_salary': result.res_mot_salary,
                    'res_mot_career': result.res_mot_career,
                    'res_mot_creativity': result.res_mot_creativity,
                    'res_mot_relationships': result.res_mot_relationships,
                    'res_mot_recognition': result.res_mot_recognition,
                    'res_mot_affiliation': result.res_mot_affiliation,
                    'res_mot_self_development': result.res_mot_self_development,
                    'res_mot_purpose': result.res_mot_purpose,
                    'res_mot_cooperation': result.res_mot_cooperation,
                    'res_mot_stability': result.res_mot_stability,
                    'res_mot_tradition': result.res_mot_tradition,
                    'res_mot_management': result.res_mot_management,
                    'res_mot_work_conditions': result.res_mot_work_conditions,
                },
                
                # Ценности
                'values': {
                    'res_val_honesty_justice': result.res_val_honesty_justice,
                    'res_val_humanism': result.res_val_humanism,
                    'res_val_patriotism': result.res_val_patriotism,
                    'res_val_family': result.res_val_family,
                    'res_val_health': result.res_val_health,
                    'res_val_environment': result.res_val_environment,
                }
            }
            results_list.append(result_data)
            
        return successResponse({
            "participant": {
                "part_id": participant.part_id,
                "part_name": participant.part_name,
                "part_gender": participant.part_gender,
                "part_institution": participant.part_institution.inst_name if participant.part_institution else None,
                "part_spec": participant.part_spec.spec_name if participant.part_spec else None,
                "part_edu_level": participant.part_edu_level.edu_level_name if participant.part_edu_level else None,
                "part_form": participant.part_form.form_name if participant.part_form else None,
                "part_course_num": participant.part_course_num
            },
            "results": results_list
        })
        
    except ValueError:
        return errorResponse("part_id must be an integer")
    except Exception as e:
        return exceptionResponse(e)


@method('POST')
@csrf_exempt
def import_excel(request):
    if request.method != "POST":
        return JsonResponse({"error": "Use POST"}, status=405)

    excel_file = request.FILES.get("file")
    if not excel_file:
        return JsonResponse({"error": "No file uploaded"}, status=400)

    config_json = request.POST.get("config_json")
    if not config_json:
        return JsonResponse({"error": "Missing config_json"}, status=400)

    try:
        config = json.loads(config_json)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON in config_json"}, status=400)

    wb = openpyxl.load_workbook(excel_file, data_only=True)

    created_count = 0
    updated_count = 0

    with transaction.atomic():
        for sheet_info in config["sheets"]:
            sheet_name = sheet_info["name"]
            start_row = sheet_info["start_row"]
            columns = sheet_info["columns"]

            if sheet_name not in wb.sheetnames:
                print(f"Лист '{sheet_name}' отсутствует — пропуск")
                continue

            ws = wb[sheet_name]

            print(f"\n=== Обработка листа: {sheet_name} ===")

            for row in ws.iter_rows(min_row=start_row):
                b_value = row[1].value
                if b_value and str(b_value).strip().lower() == "итог":
                    print(f"Достигнут итоговый ряд — стоп для {sheet_name}")
                    break

                def get_col(col_letter):
                    if not col_letter:
                        return None
                    cell = row[openpyxl.utils.column_index_from_string(col_letter) - 1]
                    return cell.value

                row_data = {k: get_col(v) for k, v in columns.items()}

                if sheet_name == "Сравнение по компетенциям":
                    center_name = clean_value(row_data.get("center_name"))
                    if not center_name:
                        continue

                    center, _ = CompetenceCenters.objects.get_or_create(center_name=center_name)
                    institution = None
                    if (inst_name := clean_value(row_data.get("inst_name"))):
                        institution, _ = Institutions.objects.get_or_create(inst_name=inst_name)
                    edu_level = None
                    if (edu_level_name := clean_value(row_data.get("edu_level_name"))):
                        edu_level, _ = Educationlevels.objects.get_or_create(edu_level_name=edu_level_name)
                    form = None
                    if (form_name := clean_value(row_data.get("form_name"))):
                        form, _ = Studyforms.objects.get_or_create(form_name=form_name)
                    spec = None
                    if (spec_name := clean_value(row_data.get("spec_name"))):
                        spec, _ = Specialties.objects.get_or_create(spec_name=spec_name)

                    participant_name = clean_value(row_data.get("part_name"))
                    if not participant_name:
                        continue

                    # ═══════════════════════════════════════════════════════════
                    # ИСПРАВЛЕНО: Уникальность по (name, institution, spec)
                    # ═══════════════════════════════════════════════════════════
                    participant, created = Participants.objects.get_or_create(
                        part_name=participant_name,
                        part_institution=institution,  # ← Теперь в условии поиска!
                        part_spec=spec,                 # ← Теперь в условии поиска!
                        defaults={
                            "part_gender": clean_value(row_data.get("part_gender")),
                            "part_edu_level": edu_level,
                            "part_form": form,
                            "part_course_num": clean_value(row_data.get("part_course_num"), "int"),
                        },
                    )

                    # Создаём или обновляем результат
                    result, created_res = Results.objects.get_or_create(
                        res_participant=participant,
                        res_year=clean_value(row_data.get("res_year")),
                        defaults={
                            "res_center": center,
                            "res_institution": institution,
                            "res_edu_level": edu_level,
                            "res_form": form,
                            "res_spec": spec,
                            "res_course_num": clean_value(row_data.get("res_course_num"), "int"),
                            "res_high_potential": clean_value(row_data.get("res_high_potential")),
                            "res_summary_report": clean_value(row_data.get("res_summary_report")),
                        },
                    )

                    # Обновляем компетенции
                    Results.objects.filter(pk=result.pk).update(
                        res_comp_info_analysis=clean_value(row_data.get("res_comp_info_analysis"), "int"),
                        res_comp_planning=clean_value(row_data.get("res_comp_planning"), "int"),
                        res_comp_result_orientation=clean_value(row_data.get("res_comp_result_orientation"), "int"),
                        res_comp_stress_resistance=clean_value(row_data.get("res_comp_stress_resistance"), "int"),
                        res_comp_partnership=clean_value(row_data.get("res_comp_partnership"), "int"),
                        res_comp_rules_compliance=clean_value(row_data.get("res_comp_rules_compliance"), "int"),
                        res_comp_self_development=clean_value(row_data.get("res_comp_self_development"), "int"),
                        res_comp_leadership=clean_value(row_data.get("res_comp_leadership"), "int"),
                        res_comp_emotional_intel=clean_value(row_data.get("res_comp_emotional_intel"), "int"),
                        res_comp_client_focus=clean_value(row_data.get("res_comp_client_focus"), "int"),
                        res_comp_communication=clean_value(row_data.get("res_comp_communication"), "int"),
                        res_comp_passive_vocab=clean_value(row_data.get("res_comp_passive_vocab"), "int"),
                    )
                    created_count += 1

                # === Остальные листы БЕЗ ИЗМЕНЕНИЙ ===
                elif sheet_name == "Мотивационный профиль":
                    participant_name = clean_value(row_data.get("part_name"))
                    if not participant_name:
                        continue

                    try:
                        # ═══════════════════════════════════════════════════════════
                        # ПРОБЛЕМА: Здесь поиск только по имени!
                        # Если у студента 2 записи (МГУ и СПбГУ), найдётся только первая
                        # ═══════════════════════════════════════════════════════════
                        participant = Participants.objects.get(part_name=participant_name)
                    except Participants.DoesNotExist:
                        print(f"Не найден участник '{participant_name}', пропуск")
                        continue
                    except Participants.MultipleObjectsReturned:
                        # Если несколько - берём первого (временное решение)
                        participant = Participants.objects.filter(part_name=participant_name).first()
                        print(f"⚠️ Найдено несколько участников с именем '{participant_name}', взят первый")

                    Results.objects.filter(
                        res_participant=participant,
                        res_year=clean_value(row_data.get("res_year"))
                    ).update(
                        res_mot_autonomy=clean_value(row_data.get("res_mot_autonomy"), "float"),
                        res_mot_altruism=clean_value(row_data.get("res_mot_altruism"), "float"),
                        res_mot_challenge=clean_value(row_data.get("res_mot_challenge"), "float"),
                        res_mot_salary=clean_value(row_data.get("res_mot_salary"), "float"),
                        res_mot_career=clean_value(row_data.get("res_mot_career"), "float"),
                        res_mot_creativity=clean_value(row_data.get("res_mot_creativity"), "float"),
                        res_mot_relationships=clean_value(row_data.get("res_mot_relationships"), "float"),
                        res_mot_recognition=clean_value(row_data.get("res_mot_recognition"), "float"),
                        res_mot_affiliation=clean_value(row_data.get("res_mot_affiliation"), "float"),
                        res_mot_self_development=clean_value(row_data.get("res_mot_self_development"), "float"),
                        res_mot_purpose=clean_value(row_data.get("res_mot_purpose"), "float"),
                        res_mot_cooperation=clean_value(row_data.get("res_mot_cooperation"), "float"),
                        res_mot_stability=clean_value(row_data.get("res_mot_stability"), "float"),
                        res_mot_tradition=clean_value(row_data.get("res_mot_tradition"), "float"),
                        res_mot_management=clean_value(row_data.get("res_mot_management"), "float"),
                        res_mot_work_conditions=clean_value(row_data.get("res_mot_work_conditions"), "float"),
                    )
                    updated_count += 1

                elif sheet_name == "Образовательные курсы":
                    participant_name = clean_value(row_data.get("part_name"))
                    if not participant_name:
                        continue
                    try:
                        participant = Participants.objects.get(part_name=participant_name)
                    except Participants.DoesNotExist:
                        continue
                    except Participants.MultipleObjectsReturned:
                        participant = Participants.objects.filter(part_name=participant_name).first()

                    Course.objects.update_or_create(
                        course_participant=participant,
                        defaults={
                            "course_an_dec": clean_value(row_data.get("course_an_dec"), "float"),
                            "course_client_focus": clean_value(row_data.get("course_client_focus"), "float"),
                            "course_communication": clean_value(row_data.get("course_communication"), "float"),
                            "course_leadership": clean_value(row_data.get("course_leadership"), "float"),
                            "course_result_orientation": clean_value(row_data.get("course_result_orientation"), "float"),
                            "course_planning_org": clean_value(row_data.get("course_planning_org"), "float"),
                            "course_rules_culture": clean_value(row_data.get("course_rules_culture"), "float"),
                            "course_self_dev": clean_value(row_data.get("course_self_dev"), "float"),
                            "course_collaboration": clean_value(row_data.get("course_collaboration"), "float"),
                            "course_stress_resistance": clean_value(row_data.get("course_stress_resistance"), "float"),
                            "course_emotions_communication": clean_value(row_data.get("course_emotions_communication"), "float"),
                            "course_negotiations": clean_value(row_data.get("course_negotiations"), "float"),
                            "course_digital_comm": clean_value(row_data.get("course_digital_comm"), "float"),
                            "course_effective_learning": clean_value(row_data.get("course_effective_learning"), "float"),
                            "course_entrepreneurship": clean_value(row_data.get("course_entrepreneurship"), "float"),
                            "course_creativity_tech": clean_value(row_data.get("course_creativity_tech"), "float"),
                            "course_trendwatching": clean_value(row_data.get("course_trendwatching"), "float"),
                            "course_conflict_management": clean_value(row_data.get("course_conflict_management"), "float"),
                            "course_career_management": clean_value(row_data.get("course_career_management"), "float"),
                            "course_burnout": clean_value(row_data.get("course_burnout"), "float"),
                            "course_cross_cultural_comm": clean_value(row_data.get("course_cross_cultural_comm"), "float"),
                            "course_mentoring": clean_value(row_data.get("course_mentoring"), "float"),
                        },
                    )
                    updated_count += 1

                elif sheet_name == "Итоги успеваемости участников":
                    participant_name = clean_value(row_data.get("part_name"))
                    if not participant_name:
                        continue

                    try:
                        participant = Participants.objects.get(part_name=participant_name)
                    except Participants.DoesNotExist:
                        print(f"Не найден участник '{participant_name}', пропуск")
                        continue
                    except Participants.MultipleObjectsReturned:
                        participant = Participants.objects.filter(part_name=participant_name).first()

                    perf_year = clean_value(row_data.get("perf_year"))
                    perf_current_avg = clean_value(row_data.get("perf_current_avg"), "float")
                    perf_digital_culture = clean_value(row_data.get("perf_digital_culture"), "float")
                    perf_main_attestation = clean_value(row_data.get("perf_main_attestation"))
                    perf_first_retake = clean_value(row_data.get("perf_first_retake"))
                    perf_second_retake = clean_value(row_data.get("perf_second_retake"))
                    perf_high_grade_retake = clean_value(row_data.get("perf_high_grade_retake"))
                    perf_final_grade = clean_value(row_data.get("perf_final_grade"))

                    Academicperformance.objects.update_or_create(
                        perf_part_id=participant.part_id,
                        perf_year=perf_year,
                        defaults={
                            "perf_current_avg": perf_current_avg,
                            "perf_digital_culture": perf_digital_culture,
                            "perf_main_attestation": perf_main_attestation,
                            "perf_first_retake": perf_first_retake,
                            "perf_second_retake": perf_second_retake,
                            "perf_high_grade_retake": perf_high_grade_retake,
                            "perf_final_grade": perf_final_grade,
                        }
                    )

                    updated_count += 1

    return JsonResponse({"status": "success", "created": created_count, "updated": updated_count})


@method('GET')
@response
@csrf_exempt
def stats(request):
    try:
        # Общая статистика
        total_participants = Participants.objects.count()
        total_tests = Results.objects.count()
        unique_institutions = Institutions.objects.count()
        unique_centers = CompetenceCenters.objects.count()

        # Участники по году получения первой оценки
        first_year_stats = Results.objects.values('res_participant').annotate(
            first_year=Min('res_year')
        ).values('first_year').annotate(
            count=Count('res_participant')
        ).order_by('first_year')
        
        participants_by_first_year = {
            'years': [str(stat['first_year']) for stat in first_year_stats if stat['first_year']],
            'counts': [stat['count'] for stat in first_year_stats if stat['first_year']]
        }

        # Участники по центрам компетенций (топ-15)
        centers_stats = Results.objects.filter(
            res_center__isnull=False
        ).values('res_center__center_name').annotate(
            count=Count('res_participant', distinct=True)
        ).order_by('-count')[:15]
        
        participants_by_center = {
            'centers': [stat['res_center__center_name'] for stat in centers_stats if stat['res_center__center_name']],
            'counts': [stat['count'] for stat in centers_stats if stat['res_center__center_name']]
        }

        # Участники по учебным заведениям (топ-15)
        institutions_stats = Results.objects.filter(
            res_institution__isnull=False
        ).values('res_institution__inst_name').annotate(
            count=Count('res_participant', distinct=True)
        ).order_by('-count')[:15]
        
        participants_by_institution = {
            'institutions': [stat['res_institution__inst_name'] for stat in institutions_stats if stat['res_institution__inst_name']],
            'counts': [stat['count'] for stat in institutions_stats if stat['res_institution__inst_name']]
        }

        # Специальности участников
        specialties_stats = Results.objects.filter(
            res_spec__isnull=False
        ).values('res_spec__spec_name').annotate(
            count=Count('res_participant', distinct=True)
        ).order_by('-count')
        
        specialties_distribution = {
            'specialties': [stat['res_spec__spec_name'] for stat in specialties_stats if stat['res_spec__spec_name']],
            'counts': [stat['count'] for stat in specialties_stats if stat['res_spec__spec_name']]
        }

        # Динамика тестирований по годам
        tests_by_year = Results.objects.filter(
            res_year__isnull=False
        ).values('res_year').annotate(
            count=Count('res_id')
        ).order_by('res_year')
        
        tests_by_year_data = {
            'years': [str(stat['res_year']) for stat in tests_by_year if stat['res_year']],
            'counts': [stat['count'] for stat in tests_by_year if stat['res_year']]
        }

        # Средние оценки по компетенциям по годам
        competences_fields = [
            'res_comp_info_analysis', 'res_comp_planning', 'res_comp_result_orientation',
            'res_comp_stress_resistance', 'res_comp_partnership', 'res_comp_rules_compliance',
            'res_comp_self_development', 'res_comp_leadership', 'res_comp_emotional_intel',
            'res_comp_client_focus', 'res_comp_communication', 'res_comp_passive_vocab'
        ]
        
        competence_names = {
            'res_comp_info_analysis': 'Анализ информации',
            'res_comp_planning': 'Планирование',
            'res_comp_result_orientation': 'Ориентация на результат',
            'res_comp_stress_resistance': 'Стрессоустойчивость',
            'res_comp_partnership': 'Партнерство',
            'res_comp_rules_compliance': 'Соблюдение правил',
            'res_comp_self_development': 'Саморазвитие',
            'res_comp_leadership': 'Лидерство',
            'res_comp_emotional_intel': 'Эмоциональный интеллект',
            'res_comp_client_focus': 'Клиентоориентированность',
            'res_comp_communication': 'Коммуникация',
            'res_comp_passive_vocab': 'Пассивный словарь'
        }
        
        competences_by_year = []
        for field in competences_fields:
            yearly_stats = Results.objects.filter(
                **{f'{field}__isnull': False},
                res_year__isnull=False
            ).values('res_year').annotate(
                avg_value=Avg(field)
            ).order_by('res_year')
            
            if yearly_stats:
                competences_by_year.append({
                    'name': competence_names[field],
                    'years': [str(stat['res_year']) for stat in yearly_stats if stat['res_year']],
                    'values': [round(float(stat['avg_value']), 1) for stat in yearly_stats if stat['res_year']]
                })

        # Средние оценки по мотиваторам по годам
        motivators_fields = [
            'res_mot_autonomy', 'res_mot_altruism', 'res_mot_challenge', 'res_mot_salary',
            'res_mot_career', 'res_mot_creativity', 'res_mot_relationships', 'res_mot_recognition',
            'res_mot_affiliation', 'res_mot_self_development', 'res_mot_purpose', 'res_mot_cooperation',
            'res_mot_stability', 'res_mot_tradition', 'res_mot_management', 'res_mot_work_conditions'
        ]
        
        motivator_names = {
            'res_mot_autonomy': 'Автономия',
            'res_mot_altruism': 'Альтруизм',
            'res_mot_challenge': 'Вызов',
            'res_mot_salary': 'Зарплата',
            'res_mot_career': 'Карьера',
            'res_mot_creativity': 'Креативность',
            'res_mot_relationships': 'Отношения',
            'res_mot_recognition': 'Признание',
            'res_mot_affiliation': 'Принадлежность',
            'res_mot_self_development': 'Саморазвитие',
            'res_mot_purpose': 'Цель',
            'res_mot_cooperation': 'Сотрудничество',
            'res_mot_stability': 'Стабильность',
            'res_mot_tradition': 'Традиции',
            'res_mot_management': 'Управление',
            'res_mot_work_conditions': 'Условия работы'
        }
        
        motivators_by_year = []
        for field in motivators_fields:
            yearly_stats = Results.objects.filter(
                **{f'{field}__isnull': False},
                res_year__isnull=False
            ).values('res_year').annotate(
                avg_value=Avg(field)
            ).order_by('res_year')
            
            if yearly_stats:
                motivators_by_year.append({
                    'name': motivator_names[field],
                    'years': [str(stat['res_year']) for stat in yearly_stats if stat['res_year']],
                    'values': [round(float(stat['avg_value']), 1) for stat in yearly_stats if stat['res_year']]
                })

        # Средние оценки по ценностям по годам
        values_fields = [
            'res_val_honesty_justice', 'res_val_humanism', 'res_val_patriotism',
            'res_val_family', 'res_val_health', 'res_val_environment'
        ]
        
        value_names = {
            'res_val_honesty_justice': 'Честность и справедливость',
            'res_val_humanism': 'Гуманизм',
            'res_val_patriotism': 'Патриотизм',
            'res_val_family': 'Семья',
            'res_val_health': 'Здоровье',
            'res_val_environment': 'Окружающая среда'
        }
        
        values_by_year = []
        for field in values_fields:
            yearly_stats = Results.objects.filter(
                **{f'{field}__isnull': False},
                res_year__isnull=False
            ).values('res_year').annotate(
                avg_value=Avg(field)
            ).order_by('res_year')
            
            if yearly_stats:
                values_by_year.append({
                    'name': value_names[field],
                    'years': [str(stat['res_year']) for stat in yearly_stats if stat['res_year']],
                    'values': [round(float(stat['avg_value']), 1) for stat in yearly_stats if stat['res_year']]
                })

        stats_data = {
            'totalParticipants': total_participants,
            'totalTests': total_tests,
            'uniqueInstitutions': unique_institutions,
            'uniqueCenters': unique_centers,
            'participantsByFirstYear': participants_by_first_year,
            'participantsByCenter': participants_by_center,
            'participantsByInstitution': participants_by_institution,
            'specialtiesDistribution': specialties_distribution,
            'testsByYear': tests_by_year_data,
            'competencesByYear': competences_by_year,
            'motivatorsByYear': motivators_by_year,
            'valuesByYear': values_by_year
        }
        
        return {"stats": stats_data}
        
    except Exception as e:
        return exceptionResponse(e)



######################## DATA VIEW SESSIONS ########################

data_view_sessions = {}  # ffs

class DataViewSession:
    def __init__(self):
        self.session_id = str(uuid.uuid4())
        self.created_at = timezone.now()
        self.last_activity = timezone.now()
        self.filters = []
        self.visible_columns = None  # None = все колонки
        self.limit = 1000
        self.offset = 0
        
    def to_dict(self):
        return {
            'session_id': self.session_id,
            'filters': self.filters,
            'visible_columns': self.visible_columns,
            'limit': self.limit,
            'offset': self.offset
        }
    
    def update_activity(self):
        self.last_activity = timezone.now()


def cleanup_old_sessions():
    now = timezone.now()
    expired_sessions = []
    for session_id, session in data_view_sessions.items():
        if (now - session.last_activity).total_seconds() > 3600:  # 1 час
            expired_sessions.append(session_id)
    
    for session_id in expired_sessions:
        del data_view_sessions[session_id]


@method('POST')
@csrf_exempt
def create_data_session(request):
    try:
        cleanup_old_sessions()  # Очищаем старые сессии
        
        session = DataViewSession()
        data_view_sessions[session.session_id] = session
        
        return successResponse({
            "session_id": session.session_id,
            "filters": session.filters,
            "visible_columns": session.visible_columns,
            "limit": session.limit
        })
        
    except Exception as e:
        return exceptionResponse(e)


@method('POST')
@csrf_exempt
def get_session_data(request):
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        
        if not session_id or session_id not in data_view_sessions:
            return errorResponse("Invalid session ID")
        
        session = data_view_sessions[session_id]
        session.update_activity()
        
        # Получаем данные с учетом фильтров и видимых колонок
        results_query = Results.objects.all().select_related(
            'res_participant', 'res_center', 'res_institution',
            'res_edu_level', 'res_form', 'res_spec'
        )
        
        # Применяем фильтры сессии
        for filter_obj in session.filters:
            if filter_obj['type'] == 'basic' and filter_obj['selectedValues']:
                field = filter_obj['field']
                if field == 'part_gender':
                    results_query = results_query.filter(
                        res_participant__part_gender__in=filter_obj['selectedValues']
                    )
                elif field == 'center':
                    results_query = results_query.filter(
                        res_center__center_name__in=filter_obj['selectedValues']
                    )
                elif field == 'institution':
                    results_query = results_query.filter(
                        res_institution__inst_name__in=filter_obj['selectedValues']
                    )
                elif field == 'edu_level':
                    results_query = results_query.filter(
                        res_edu_level__edu_level_name__in=filter_obj['selectedValues']
                    )
                elif field == 'study_form':
                    results_query = results_query.filter(
                        res_form__form_name__in=filter_obj['selectedValues']
                    )
                elif field == 'specialty':
                    results_query = results_query.filter(
                        res_spec__spec_name__in=filter_obj['selectedValues']
                    )
                elif field == 'res_year':
                    results_query = results_query.filter(
                        res_year__in=filter_obj['selectedValues']
                    )
                elif field == 'res_course_num':
                    results_query = results_query.filter(
                        res_course_num__in=filter_obj['selectedValues']
                    )
                    
            elif filter_obj['type'] == 'numeric':
                field = filter_obj['field']
                min_val = filter_obj['min']
                max_val = filter_obj['max']
                
                # Компетенции
                if field.startswith('res_comp_'):
                    results_query = results_query.filter(
                        **{f'{field}__gte': min_val, f'{field}__lte': max_val}
                    )
                # Мотиваторы
                elif field.startswith('res_mot_'):
                    results_query = results_query.filter(
                        **{f'{field}__gte': min_val, f'{field}__lte': max_val}
                    )
                # Ценности
                elif field.startswith('res_val_'):
                    results_query = results_query.filter(
                        **{f'{field}__gte': min_val, f'{field}__lte': max_val}
                    )
        
        total_count = results_query.count()
        
        # Применяем лимит и оффсет
        results_list = results_query[session.offset:session.offset + session.limit]
        
        # Формируем данные
        results_data = []
        for result in results_list:
            result_data = format_result_data(result, session.visible_columns)
            results_data.append(result_data)
        
        return successResponse({
            "results": results_data,
            "total_count": total_count,
            "session": session.to_dict()
        })
        
    except Exception as e:
        return exceptionResponse(e)


@method('POST')
@csrf_exempt
def update_session_filters(request):
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        filters = data.get('filters', [])
        
        if not session_id or session_id not in data_view_sessions:
            return errorResponse("Invalid session ID")
        
        session = data_view_sessions[session_id]
        session.update_activity()
        session.filters = filters
        session.offset = 0  # Сбрасываем оффсет при изменении фильтров
        
        return successResponse({
            "session_id": session.session_id,
            "filters": session.filters,
            "message": "Filters updated successfully"
        })
        
    except Exception as e:
        return exceptionResponse(e)


@method('POST')
@csrf_exempt
def update_session_columns(request):
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        visible_columns = data.get('visible_columns')
        
        if not session_id or session_id not in data_view_sessions:
            return errorResponse("Invalid session ID")
        
        session = data_view_sessions[session_id]
        session.update_activity()
        session.visible_columns = visible_columns
        
        return successResponse({
            "session_id": session.session_id,
            "visible_columns": session.visible_columns,
            "message": "Visible columns updated successfully"
        })
        
    except Exception as e:
        return exceptionResponse(e)


@method('POST')
@csrf_exempt
def load_more_data(request):
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        
        if not session_id or session_id not in data_view_sessions:
            return errorResponse("Invalid session ID")
        
        session = data_view_sessions[session_id]
        session.update_activity()
        session.offset += session.limit
        
        # Повторно используем логику get_session_data
        return get_session_data(request)
        
    except Exception as e:
        return exceptionResponse(e)


@method('POST')
@csrf_exempt
def export_session_data(request):
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        
        if not session_id or session_id not in data_view_sessions:
            return errorResponse("Invalid session ID")
        
        session = data_view_sessions[session_id]
        session.update_activity()
        
        # Получаем ВСЕ данные с текущими фильтрами (без лимита)
        results_query = Results.objects.all().select_related(
            'res_participant', 'res_center', 'res_institution',
            'res_edu_level', 'res_form', 'res_spec'
        )
        
        # Применяем фильтры сессии
        for filter_obj in session.filters:
            if filter_obj['type'] == 'basic' and filter_obj['selectedValues']:
                field = filter_obj['field']
                if field == 'part_gender':
                    results_query = results_query.filter(
                        res_participant__part_gender__in=filter_obj['selectedValues']
                    )
                elif field == 'center':
                    results_query = results_query.filter(
                        res_center__center_name__in=filter_obj['selectedValues']
                    )
                elif field == 'institution':
                    results_query = results_query.filter(
                        res_institution__inst_name__in=filter_obj['selectedValues']
                    )
                elif field == 'edu_level':
                    results_query = results_query.filter(
                        res_edu_level__edu_level_name__in=filter_obj['selectedValues']
                    )
                elif field == 'study_form':
                    results_query = results_query.filter(
                        res_form__form_name__in=filter_obj['selectedValues']
                    )
                elif field == 'specialty':
                    results_query = results_query.filter(
                        res_spec__spec_name__in=filter_obj['selectedValues']
                    )
                elif field == 'res_year':
                    results_query = results_query.filter(
                        res_year__in=filter_obj['selectedValues']
                    )
                elif field == 'res_course_num':
                    results_query = results_query.filter(
                        res_course_num__in=filter_obj['selectedValues']
                    )
                    
            elif filter_obj['type'] == 'numeric':
                field = filter_obj['field']
                min_val = filter_obj['min']
                max_val = filter_obj['max']
                
                # Компетенции
                if field.startswith('res_comp_'):
                    results_query = results_query.filter(
                        **{f'{field}__gte': min_val, f'{field}__lte': max_val}
                    )
                # Мотиваторы
                elif field.startswith('res_mot_'):
                    results_query = results_query.filter(
                        **{f'{field}__gte': min_val, f'{field}__lte': max_val}
                    )
                # Ценности
                elif field.startswith('res_val_'):
                    results_query = results_query.filter(
                        **{f'{field}__gte': min_val, f'{field}__lte': max_val}
                    )
        
        results_list = results_query
        
        # Создаем DataFrame для экспорта
        export_data = []
        for result in results_list:
            row = format_result_for_export(result, session.visible_columns)
            export_data.append(row)
        
        df = pd.DataFrame(export_data)
        
        # Создаем HTTP response с Excel файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="results_export.xlsx"'
        
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Результаты тестирования')
        
        return response
        
    except Exception as e:
        return exceptionResponse(e)


@method('POST')
@csrf_exempt
def export_selected_results(request):
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        selected_ids = data.get('selected_ids', [])
        
        if not session_id or session_id not in data_view_sessions:
            return errorResponse("Invalid session ID")
        
        if not selected_ids:
            return errorResponse("No records selected for export")
        
        session = data_view_sessions[session_id]
        session.update_activity()
        
        # Получаем выбранные записи
        results_query = Results.objects.filter(
            res_id__in=selected_ids
        ).select_related(
            'res_participant', 'res_center', 'res_institution',
            'res_edu_level', 'res_form', 'res_spec'
        )
        
        # Создаем DataFrame для экспорта
        export_data = []
        for result in results_query:
            row = format_result_for_export(result, session.visible_columns)
            export_data.append(row)
        
        df = pd.DataFrame(export_data)
        
        # Создаем HTTP response с Excel файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="selected_results_export.xlsx"'
        
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Выбранные результаты')
        
        return response
        
    except Exception as e:
        return exceptionResponse(e)


@method('POST')
@csrf_exempt
def stats_with_filters(request):
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        
        # Если есть session_id, используем фильтры из сессии
        filters = []
        if session_id and session_id in data_view_sessions:
            session = data_view_sessions[session_id]
            session.update_activity()
            filters = session.filters
        
        # Общая статистика
        total_participants = Participants.objects.count()
        total_tests = Results.objects.count()
        unique_institutions = Institutions.objects.count()
        unique_centers = CompetenceCenters.objects.count()

        # Участники по году получения первой оценки
        first_year_stats = Results.objects.values('res_participant').annotate(
            first_year=Min('res_year')
        ).values('first_year').annotate(
            count=Count('res_participant')
        ).order_by('first_year')
        
        participants_by_first_year = {
            'years': [str(stat['first_year']) for stat in first_year_stats if stat['first_year']],
            'counts': [stat['count'] for stat in first_year_stats if stat['first_year']]
        }

        # Участники по центрам компетенций (топ-15)
        centers_stats = Results.objects.filter(
            res_center__isnull=False
        ).values('res_center__center_name').annotate(
            count=Count('res_participant', distinct=True)
        ).order_by('-count')[:15]
        
        participants_by_center = {
            'centers': [stat['res_center__center_name'] for stat in centers_stats if stat['res_center__center_name']],
            'counts': [stat['count'] for stat in centers_stats if stat['res_center__center_name']]
        }

        # Участники по учебным заведениям (топ-15)
        institutions_stats = Results.objects.filter(
            res_institution__isnull=False
        ).values('res_institution__inst_name').annotate(
            count=Count('res_participant', distinct=True)
        ).order_by('-count')[:15]
        
        participants_by_institution = {
            'institutions': [stat['res_institution__inst_name'] for stat in institutions_stats if stat['res_institution__inst_name']],
            'counts': [stat['count'] for stat in institutions_stats if stat['res_institution__inst_name']]
        }

        # Специальности участников
        specialties_stats = Results.objects.filter(
            res_spec__isnull=False
        ).values('res_spec__spec_name').annotate(
            count=Count('res_participant', distinct=True)
        ).order_by('-count')
        
        specialties_distribution = {
            'specialties': [stat['res_spec__spec_name'] for stat in specialties_stats if stat['res_spec__spec_name']],
            'counts': [stat['count'] for stat in specialties_stats if stat['res_spec__spec_name']]
        }

        # Динамика тестирований по годам
        tests_by_year = Results.objects.filter(
            res_year__isnull=False
        ).values('res_year').annotate(
            count=Count('res_id')
        ).order_by('res_year')
        
        tests_by_year_data = {
            'years': [str(stat['res_year']) for stat in tests_by_year if stat['res_year']],
            'counts': [stat['count'] for stat in tests_by_year if stat['res_year']]
        }

        # Средние оценки по компетенциям по годам
        competences_fields = [
            'res_comp_info_analysis', 'res_comp_planning', 'res_comp_result_orientation',
            'res_comp_stress_resistance', 'res_comp_partnership', 'res_comp_rules_compliance',
            'res_comp_self_development', 'res_comp_leadership', 'res_comp_emotional_intel',
            'res_comp_client_focus', 'res_comp_communication', 'res_comp_passive_vocab'
        ]
        
        competence_names = {
            'res_comp_info_analysis': 'Анализ информации',
            'res_comp_planning': 'Планирование',
            'res_comp_result_orientation': 'Ориентация на результат',
            'res_comp_stress_resistance': 'Стрессоустойчивость',
            'res_comp_partnership': 'Партнерство',
            'res_comp_rules_compliance': 'Соблюдение правил',
            'res_comp_self_development': 'Саморазвитие',
            'res_comp_leadership': 'Лидерство',
            'res_comp_emotional_intel': 'Эмоциональный интеллект',
            'res_comp_client_focus': 'Клиентоориентированность',
            'res_comp_communication': 'Коммуникация',
            'res_comp_passive_vocab': 'Пассивный словарь'
        }
        
        competences_by_year = []
        for field in competences_fields:
            yearly_stats = Results.objects.filter(
                **{f'{field}__isnull': False},
                res_year__isnull=False
            ).values('res_year').annotate(
                avg_value=Avg(field)
            ).order_by('res_year')
            
            if yearly_stats:
                competences_by_year.append({
                    'name': competence_names[field],
                    'years': [str(stat['res_year']) for stat in yearly_stats if stat['res_year']],
                    'values': [round(float(stat['avg_value']), 1) for stat in yearly_stats if stat['res_year']]
                })

        # Средние оценки по мотиваторам по годам
        motivators_fields = [
            'res_mot_autonomy', 'res_mot_altruism', 'res_mot_challenge', 'res_mot_salary',
            'res_mot_career', 'res_mot_creativity', 'res_mot_relationships', 'res_mot_recognition',
            'res_mot_affiliation', 'res_mot_self_development', 'res_mot_purpose', 'res_mot_cooperation',
            'res_mot_stability', 'res_mot_tradition', 'res_mot_management', 'res_mot_work_conditions'
        ]
        
        motivator_names = {
            'res_mot_autonomy': 'Автономия',
            'res_mot_altruism': 'Альтруизм',
            'res_mot_challenge': 'Вызов',
            'res_mot_salary': 'Зарплата',
            'res_mot_career': 'Карьера',
            'res_mot_creativity': 'Креативность',
            'res_mot_relationships': 'Отношения',
            'res_mot_recognition': 'Признание',
            'res_mot_affiliation': 'Принадлежность',
            'res_mot_self_development': 'Саморазвитие',
            'res_mot_purpose': 'Цель',
            'res_mot_cooperation': 'Сотрудничество',
            'res_mot_stability': 'Стабильность',
            'res_mot_tradition': 'Традиции',
            'res_mot_management': 'Управление',
            'res_mot_work_conditions': 'Условия работы'
        }
        
        motivators_by_year = []
        for field in motivators_fields:
            yearly_stats = Results.objects.filter(
                **{f'{field}__isnull': False},
                res_year__isnull=False
            ).values('res_year').annotate(
                avg_value=Avg(field)
            ).order_by('res_year')
            
            if yearly_stats:
                motivators_by_year.append({
                    'name': motivator_names[field],
                    'years': [str(stat['res_year']) for stat in yearly_stats if stat['res_year']],
                    'values': [round(float(stat['avg_value']), 1) for stat in yearly_stats if stat['res_year']]
                })

        # Средние оценки по ценностям по годам
        values_fields = [
            'res_val_honesty_justice', 'res_val_humanism', 'res_val_patriotism',
            'res_val_family', 'res_val_health', 'res_val_environment'
        ]
        
        value_names = {
            'res_val_honesty_justice': 'Честность и справедливость',
            'res_val_humanism': 'Гуманизм',
            'res_val_patriotism': 'Патриотизм',
            'res_val_family': 'Семья',
            'res_val_health': 'Здоровье',
            'res_val_environment': 'Окружающая среда'
        }
        
        values_by_year = []
        for field in values_fields:
            yearly_stats = Results.objects.filter(
                **{f'{field}__isnull': False},
                res_year__isnull=False
            ).values('res_year').annotate(
                avg_value=Avg(field)
            ).order_by('res_year')
            
            if yearly_stats:
                values_by_year.append({
                    'name': value_names[field],
                    'years': [str(stat['res_year']) for stat in yearly_stats if stat['res_year']],
                    'values': [round(float(stat['avg_value']), 1) for stat in yearly_stats if stat['res_year']]
                })
        
        # Добавляем available_values для фильтрации
        stats_data = {
            'totalParticipants': total_participants,
            'totalTests': total_tests,
            'uniqueInstitutions': unique_institutions,
            'uniqueCenters': unique_centers,
            'participantsByFirstYear': participants_by_first_year,
            'participantsByCenter': participants_by_center,
            'participantsByInstitution': participants_by_institution,
            'specialtiesDistribution': specialties_distribution,
            'testsByYear': tests_by_year_data,
            'competencesByYear': competences_by_year,
            'motivatorsByYear': motivators_by_year,
            'valuesByYear': values_by_year,
            'available_values': extract_available_values_for_filters(filters)
        }
        
        return successResponse({"stats": stats_data})
        
    except Exception as e:
        return exceptionResponse(e)


@method('POST')
@csrf_exempt
def group_data(request):
    try:
        data = json.loads(request.body)
        selected_ids = data.get('selected_ids', [])
        grouping_column = data.get('grouping_column')
        
        if not selected_ids:
            return errorResponse("No records selected for grouping")
        if not grouping_column:
            return errorResponse("Grouping column not specified")
        
        # Получаем выбранные записи
        results_query = Results.objects.filter(
            res_id__in=selected_ids
        ).select_related(
            'res_participant', 'res_center', 'res_institution',
            'res_edu_level', 'res_form', 'res_spec'
        )
        
        # Группируем данные
        grouped_data = {
            'competences': {},
            'motivators': {},
            'values': {}
        }
        
        # Получаем уникальные группы
        groups = set()
        for result in results_query:
            group_value = get_group_value(result, grouping_column)
            if group_value:
                groups.add(group_value)
        
        groups = sorted(list(groups))
        
        # Компетенции
        competence_fields = ['res_comp_info_analysis', 'res_comp_planning', 'res_comp_result_orientation',
                            'res_comp_stress_resistance', 'res_comp_partnership', 'res_comp_rules_compliance',
                            'res_comp_self_development', 'res_comp_leadership', 'res_comp_emotional_intel',
                            'res_comp_client_focus', 'res_comp_communication', 'res_comp_passive_vocab']
        
        for field in competence_fields:
            values_by_group = []
            for group in groups:
                group_results = [r for r in results_query if get_group_value(r, grouping_column) == group]
                field_values = [getattr(r, field) for r in group_results if getattr(r, field) is not None]
                avg_value = sum(field_values) / len(field_values) if field_values else 0
                values_by_group.append(round(avg_value, 1))
            
            grouped_data['competences'][field] = {
                'groups': groups,
                'values': values_by_group
            }
        
        # Мотиваторы
        motivator_fields = ['res_mot_autonomy', 'res_mot_altruism', 'res_mot_challenge', 'res_mot_salary',
                           'res_mot_career', 'res_mot_creativity', 'res_mot_relationships', 'res_mot_recognition',
                           'res_mot_affiliation', 'res_mot_self_development', 'res_mot_purpose', 'res_mot_cooperation',
                           'res_mot_stability', 'res_mot_tradition', 'res_mot_management', 'res_mot_work_conditions']
        
        for field in motivator_fields:
            values_by_group = []
            for group in groups:
                group_results = [r for r in results_query if get_group_value(r, grouping_column) == group]
                field_values = [getattr(r, field) for r in group_results if getattr(r, field) is not None]
                avg_value = sum(field_values) / len(field_values) if field_values else 0
                values_by_group.append(round(avg_value, 1))
            
            grouped_data['motivators'][field] = {
                'groups': groups,
                'values': values_by_group
            }
        
        # Ценности
        value_fields = ['res_val_honesty_justice', 'res_val_humanism', 'res_val_patriotism',
                       'res_val_family', 'res_val_health', 'res_val_environment']
        
        for field in value_fields:
            values_by_group = []
            for group in groups:
                group_results = [r for r in results_query if get_group_value(r, grouping_column) == group]
                field_values = [getattr(r, field) for r in group_results if getattr(r, field) is not None]
                avg_value = sum(field_values) / len(field_values) if field_values else 0
                values_by_group.append(round(avg_value, 1))
            
            grouped_data['values'][field] = {
                'groups': groups,
                'values': values_by_group
            }
        
        return successResponse({
            "grouped_data": grouped_data,
            "groups": groups,
            "total_records": len(selected_ids)
        })
        
    except Exception as e:
        return exceptionResponse(e)


@method('GET')
@response
@csrf_exempt
def courses(request):
    try:
        # Получаем данные по курсам с информацией об участниках
        courses_query = Course.objects.all().select_related(
            'course_participant',
            'course_participant__part_institution',
            'course_participant__part_spec',
            'course_participant__part_edu_level',
            'course_participant__part_form'
        )
        
        courses_data = []
        for course in courses_query:
            course_data = {
                'course_id': course.course_id,
                'participant': {
                    'part_id': course.course_participant.part_id,
                    'part_name': course.course_participant.part_name,
                    'part_gender': course.course_participant.part_gender,
                    'institution': course.course_participant.part_institution.inst_name if course.course_participant.part_institution else None,
                    'specialty': course.course_participant.part_spec.spec_name if course.course_participant.part_spec else None,
                    'edu_level': course.course_participant.part_edu_level.edu_level_name if course.course_participant.part_edu_level else None,
                    'study_form': course.course_participant.part_form.form_name if course.course_participant.part_form else None,
                },
                # Данные по курсам
                'course_an_dec': float(course.course_an_dec) if course.course_an_dec is not None else 0,
                'course_client_focus': float(course.course_client_focus) if course.course_client_focus is not None else 0,
                'course_communication': float(course.course_communication) if course.course_communication is not None else 0,
                'course_leadership': float(course.course_leadership) if course.course_leadership is not None else 0,
                'course_result_orientation': float(course.course_result_orientation) if course.course_result_orientation is not None else 0,
                'course_planning_org': float(course.course_planning_org) if course.course_planning_org is not None else 0,
                'course_rules_culture': float(course.course_rules_culture) if course.course_rules_culture is not None else 0,
                'course_self_dev': float(course.course_self_dev) if course.course_self_dev is not None else 0,
                'course_collaboration': float(course.course_collaboration) if course.course_collaboration is not None else 0,
                'course_stress_resistance': float(course.course_stress_resistance) if course.course_stress_resistance is not None else 0,
                'course_emotions_communication': float(course.course_emotions_communication) if course.course_emotions_communication is not None else 0,
                'course_negotiations': float(course.course_negotiations) if course.course_negotiations is not None else 0,
                'course_digital_comm': float(course.course_digital_comm) if course.course_digital_comm is not None else 0,
                'course_effective_learning': float(course.course_effective_learning) if course.course_effective_learning is not None else 0,
                'course_entrepreneurship': float(course.course_entrepreneurship) if course.course_entrepreneurship is not None else 0,
                'course_creativity_tech': float(course.course_creativity_tech) if course.course_creativity_tech is not None else 0,
                'course_trendwatching': float(course.course_trendwatching) if course.course_trendwatching is not None else 0,
                'course_conflict_management': float(course.course_conflict_management) if course.course_conflict_management is not None else 0,
                'course_career_management': float(course.course_career_management) if course.course_career_management is not None else 0,
                'course_burnout': float(course.course_burnout) if course.course_burnout is not None else 0,
                'course_cross_cultural_comm': float(course.course_cross_cultural_comm) if course.course_cross_cultural_comm is not None else 0,
                'course_mentoring': float(course.course_mentoring) if course.course_mentoring is not None else 0,
            }
            courses_data.append(course_data)
        
        return {"courses": courses_data}
        
    except Exception as e:
        return exceptionResponse(e)


def get_group_value(result, grouping_column):
    """Получает значение для группировки из результата"""
    if grouping_column == 'part_gender':
        return result.res_participant.part_gender if result.res_participant else None
    elif grouping_column == 'center':
        return result.res_center.center_name if result.res_center else None
    elif grouping_column == 'institution':
        return result.res_institution.inst_name if result.res_institution else None
    elif grouping_column == 'edu_level':
        return result.res_edu_level.edu_level_name if result.res_edu_level else None
    elif grouping_column == 'study_form':
        return result.res_form.form_name if result.res_form else None
    elif grouping_column == 'specialty':
        return result.res_spec.spec_name if result.res_spec else None
    elif grouping_column == 'res_year':
        return result.res_year
    elif grouping_column == 'res_course_num':
        return result.res_course_num
    return None


def extract_available_values_for_filters(current_filters):
    """Извлекает доступные значения для фильтрации с учетом текущих фильтров"""
    values = {}
    
    # Базовые поля для фильтрации
    basic_fields = [
        'res_year', 'part_gender', 'center', 'institution', 
        'edu_level', 'res_course_num', 'study_form', 'specialty'
    ]
    
    # Создаем базовый запрос
    results_query = Results.objects.all()
    
    # Применяем текущие фильтры (кроме того, для которого извлекаем значения)
    for filter_obj in current_filters:
        if filter_obj['type'] == 'basic' and filter_obj['selectedValues']:
            field = filter_obj['field']
            if field == 'part_gender':
                results_query = results_query.filter(
                    res_participant__part_gender__in=filter_obj['selectedValues']
                )
            elif field == 'center':
                results_query = results_query.filter(
                    res_center__center_name__in=filter_obj['selectedValues']
                )
            elif field == 'institution':
                results_query = results_query.filter(
                    res_institution__inst_name__in=filter_obj['selectedValues']
                )
            elif field == 'edu_level':
                results_query = results_query.filter(
                    res_edu_level__edu_level_name__in=filter_obj['selectedValues']
                )
            elif field == 'study_form':
                results_query = results_query.filter(
                    res_form__form_name__in=filter_obj['selectedValues']
                )
            elif field == 'specialty':
                results_query = results_query.filter(
                    res_spec__spec_name__in=filter_obj['selectedValues']
                )
            elif field == 'res_year':
                results_query = results_query.filter(
                    res_year__in=filter_obj['selectedValues']
                )
            elif field == 'res_course_num':
                results_query = results_query.filter(
                    res_course_num__in=filter_obj['selectedValues']
                )
    
    # Извлекаем уникальные значения для каждого поля
    for field in basic_fields:
        if field == 'part_gender':
            unique_values = results_query.filter(
                res_participant__part_gender__isnull=False
            ).values_list('res_participant__part_gender', flat=True).distinct().order_by('res_participant__part_gender')
        elif field == 'center':
            unique_values = results_query.filter(
                res_center__isnull=False
            ).values_list('res_center__center_name', flat=True).distinct().order_by('res_center__center_name')
        elif field == 'institution':
            unique_values = results_query.filter(
                res_institution__isnull=False
            ).values_list('res_institution__inst_name', flat=True).distinct().order_by('res_institution__inst_name')
        elif field == 'edu_level':
            unique_values = results_query.filter(
                res_edu_level__isnull=False
            ).values_list('res_edu_level__edu_level_name', flat=True).distinct().order_by('res_edu_level__edu_level_name')
        elif field == 'study_form':
            unique_values = results_query.filter(
                res_form__isnull=False
            ).values_list('res_form__form_name', flat=True).distinct().order_by('res_form__form_name')
        elif field == 'specialty':
            unique_values = results_query.filter(
                res_spec__isnull=False
            ).values_list('res_spec__spec_name', flat=True).distinct().order_by('res_spec__spec_name')
        elif field == 'res_year':
            unique_values = results_query.filter(
                res_year__isnull=False
            ).values_list('res_year', flat=True).distinct().order_by('res_year')
        elif field == 'res_course_num':
            unique_values = results_query.filter(
                res_course_num__isnull=False
            ).values_list('res_course_num', flat=True).distinct().order_by('res_course_num')
        
        # Преобразуем в список строк и фильтруем пустые значения
        values[field] = [str(val) for val in unique_values if val is not None and str(val).strip() != '']
    
    return values


def format_result_data(result, visible_columns=None):
    """Форматирует данные результата для отправки на фронтенд"""
    base_data = {
        'res_id': result.res_id,
        'participant': {
            'part_id': result.res_participant.part_id,
            'part_name': result.res_participant.part_name,
            'part_gender': result.res_participant.part_gender,
        },
        'center': result.res_center.center_name if result.res_center else None,
        'institution': result.res_institution.inst_name if result.res_institution else None,
        'edu_level': result.res_edu_level.edu_level_name if result.res_edu_level else None,
        'study_form': result.res_form.form_name if result.res_form else None,
        'specialty': result.res_spec.spec_name if result.res_spec else None,
        'res_year': result.res_year,
        'res_course_num': result.res_course_num,
        'res_high_potential': result.res_high_potential,
        'res_summary_report': result.res_summary_report,
    }
    
    # Добавляем компетенции, мотиваторы и ценности
    competence_fields = ['res_comp_info_analysis', 'res_comp_planning', 'res_comp_result_orientation',
                        'res_comp_stress_resistance', 'res_comp_partnership', 'res_comp_rules_compliance',
                        'res_comp_self_development', 'res_comp_leadership', 'res_comp_emotional_intel',
                        'res_comp_client_focus', 'res_comp_communication', 'res_comp_passive_vocab']
    
    motivator_fields = ['res_mot_autonomy', 'res_mot_altruism', 'res_mot_challenge', 'res_mot_salary',
                       'res_mot_career', 'res_mot_creativity', 'res_mot_relationships', 'res_mot_recognition',
                       'res_mot_affiliation', 'res_mot_self_development', 'res_mot_purpose', 'res_mot_cooperation',
                       'res_mot_stability', 'res_mot_tradition', 'res_mot_management', 'res_mot_work_conditions']
    
    value_fields = ['res_val_honesty_justice', 'res_val_humanism', 'res_val_patriotism',
                   'res_val_family', 'res_val_health', 'res_val_environment']
    
    # Если указаны видимые колонки, фильтруем данные
    if visible_columns:
        filtered_data = {}
        for field in visible_columns:
            if field in base_data:
                filtered_data[field] = base_data[field]
            elif field in competence_fields:
                if 'competences' not in filtered_data:
                    filtered_data['competences'] = {}
                filtered_data['competences'][field] = getattr(result, field)
            elif field in motivator_fields:
                if 'motivators' not in filtered_data:
                    filtered_data['motivators'] = {}
                filtered_data['motivators'][field] = getattr(result, field)
            elif field in value_fields:
                if 'values' not in filtered_data:
                    filtered_data['values'] = {}
                filtered_data['values'][field] = getattr(result, field)
        return filtered_data
    else:
        # Возвращаем все данные
        base_data['competences'] = {field: getattr(result, field) for field in competence_fields}
        base_data['motivators'] = {field: getattr(result, field) for field in motivator_fields}
        base_data['values'] = {field: getattr(result, field) for field in value_fields}
        return base_data


def format_result_for_export(result, visible_columns=None):
    """Форматирует данные для экспорта в Excel"""
    # Аналогичная логика format_result_data, но в плоском формате для Excel
    row = {
        'ID результата': result.res_id,
        'ФИО участника': result.res_participant.part_name,
        'Пол': result.res_participant.part_gender,
        'Центр компетенций': result.res_center.center_name if result.res_center else '',
        'Учебное заведение': result.res_institution.inst_name if result.res_institution else '',
        'Уровень образования': result.res_edu_level.edu_level_name if result.res_edu_level else '',
        'Форма обучения': result.res_form.form_name if result.res_form else '',
        'Специальность': result.res_spec.spec_name if result.res_spec else '',
        'Учебный год': result.res_year,
        'Номер курса': result.res_course_num,
        'Высокий потенциал': result.res_high_potential or '',
        'Сводный отчет': result.res_summary_report or '',
    }
    
    # Добавляем все поля компетенций, мотиваторов и ценностей
    competence_fields = ['res_comp_info_analysis', 'res_comp_planning', 'res_comp_result_orientation',
                        'res_comp_stress_resistance', 'res_comp_partnership', 'res_comp_rules_compliance',
                        'res_comp_self_development', 'res_comp_leadership', 'res_comp_emotional_intel',
                        'res_comp_client_focus', 'res_comp_communication', 'res_comp_passive_vocab']
    
    motivator_fields = ['res_mot_autonomy', 'res_mot_altruism', 'res_mot_challenge', 'res_mot_salary',
                       'res_mot_career', 'res_mot_creativity', 'res_mot_relationships', 'res_mot_recognition',
                       'res_mot_affiliation', 'res_mot_self_development', 'res_mot_purpose', 'res_mot_cooperation',
                       'res_mot_stability', 'res_mot_tradition', 'res_mot_management', 'res_mot_work_conditions']
    
    value_fields = ['res_val_honesty_justice', 'res_val_humanism', 'res_val_patriotism',
                   'res_val_family', 'res_val_health', 'res_val_environment']
    
    field_names = {
        # Компетенции
        'res_comp_info_analysis': 'Анализ информации',
        'res_comp_planning': 'Планирование',
        'res_comp_result_orientation': 'Ориентация на результат',
        'res_comp_stress_resistance': 'Стрессоустойчивость',
        'res_comp_partnership': 'Партнерство',
        'res_comp_rules_compliance': 'Соблюдение правил',
        'res_comp_self_development': 'Саморазвитие',
        'res_comp_leadership': 'Лидерство',
        'res_comp_emotional_intel': 'Эмоциональный интеллект',
        'res_comp_client_focus': 'Клиентоориентированность',
        'res_comp_communication': 'Коммуникация',
        'res_comp_passive_vocab': 'Пассивный словарь',
        
        # Мотиваторы
        'res_mot_autonomy': 'Автономия',
        'res_mot_altruism': 'Альтруизм',
        'res_mot_challenge': 'Вызов',
        'res_mot_salary': 'Зарплата',
        'res_mot_career': 'Карьера',
        'res_mot_creativity': 'Креативность',
        'res_mot_relationships': 'Отношения',
        'res_mot_recognition': 'Признание',
        'res_mot_affiliation': 'Принадлежность',
        'res_mot_self_development': 'Саморазвитие (мотиватор)',
        'res_mot_purpose': 'Цель',
        'res_mot_cooperation': 'Сотрудничество',
        'res_mot_stability': 'Стабильность',
        'res_mot_tradition': 'Традиции',
        'res_mot_management': 'Управление',
        'res_mot_work_conditions': 'Условия работы',
        
        # Ценности
        'res_val_honesty_justice': 'Честность и справедливость',
        'res_val_humanism': 'Гуманизм',
        'res_val_patriotism': 'Патриотизм',
        'res_val_family': 'Семья',
        'res_val_health': 'Здоровье',
        'res_val_environment': 'Окружающая среда'
    }
    
    # Добавляем поля в row
    all_fields = competence_fields + motivator_fields + value_fields
    for field in all_fields:
        if not visible_columns or field in visible_columns:
            value = getattr(result, field)
            row[field_names[field]] = value if value is not None else ''
    
    return row


@method('GET')
@csrf_exempt
def students(request):
    """Получить список всех студентов"""
    try:
        students_list = Participants.objects.all().select_related(
            'part_institution', 'part_spec', 'part_edu_level', 'part_form'
        )
        
        students_data = []
        for student in students_list:
            student_data = {
                'stud_id': student.part_id,
                'stud_name': student.part_name,
                'stud_gender': student.part_gender,
                'institution': student.part_institution.inst_name if student.part_institution else None,
                'specialty': student.part_spec.spec_name if student.part_spec else None,
                'edu_level': student.part_edu_level.edu_level_name if student.part_edu_level else None,
                'study_form': student.part_form.form_name if student.part_form else None,
                'course_num': student.part_course_num
            }
            students_data.append(student_data)
            
        return successResponse({
            "students": students_data,
            "total_count": len(students_data)
        })
        
    except Exception as e:
        return exceptionResponse(e)


@method('GET')
@csrf_exempt
def student_results(request):
    """Получить результаты конкретного студента"""
    stud_id = request.GET.get('stud_id')
    
    if not stud_id:
        return errorResponse("stud_id parameter is required")
    
    try:
        stud_id = int(stud_id)
        
        try:
            # Ищем участника по part_id
            participant = Participants.objects.select_related(
                'part_institution', 'part_spec', 'part_edu_level', 'part_form'
            ).get(part_id=stud_id)
        except Participants.DoesNotExist:
            return notFoundResponse("Student not found")
        
        results_list = []
        for result in Results.objects.filter(res_participant=stud_id).select_related(
            'res_center', 'res_institution', 'res_edu_level', 'res_form', 'res_spec'
        ):
            # Создаем плоскую структуру
            result_data = {
                'res_id': result.res_id,
                'res_year': result.res_year,
                'res_course_num': result.res_course_num,
                'res_high_potential': result.res_high_potential,
                'res_summary_report': result.res_summary_report,
                
                # Базовые данные
                'center': result.res_center.center_name if result.res_center else None,
                'institution': result.res_institution.inst_name if result.res_institution else None,
                'edu_level': result.res_edu_level.edu_level_name if result.res_edu_level else None,
                'study_form': result.res_form.form_name if result.res_form else None,
                'specialty': result.res_spec.spec_name if result.res_spec else None,
                
                # Компетенции (разворачиваем из объекта competences)
                'res_comp_info_analysis': result.res_comp_info_analysis,
                'res_comp_planning': result.res_comp_planning,
                'res_comp_result_orientation': result.res_comp_result_orientation,
                'res_comp_stress_resistance': result.res_comp_stress_resistance,
                'res_comp_partnership': result.res_comp_partnership,
                'res_comp_rules_compliance': result.res_comp_rules_compliance,
                'res_comp_self_development': result.res_comp_self_development,
                'res_comp_leadership': result.res_comp_leadership,
                'res_comp_emotional_intel': result.res_comp_emotional_intel,
                'res_comp_client_focus': result.res_comp_client_focus,
                'res_comp_communication': result.res_comp_communication,
                'res_comp_passive_vocab': result.res_comp_passive_vocab,
                
                # Мотиваторы (разворачиваем из объекта motivators)
                'res_mot_autonomy': float(result.res_mot_autonomy) if result.res_mot_autonomy else None,
                'res_mot_altruism': float(result.res_mot_altruism) if result.res_mot_altruism else None,
                'res_mot_challenge': float(result.res_mot_challenge) if result.res_mot_challenge else None,
                'res_mot_salary': float(result.res_mot_salary) if result.res_mot_salary else None,
                'res_mot_career': float(result.res_mot_career) if result.res_mot_career else None,
                'res_mot_creativity': float(result.res_mot_creativity) if result.res_mot_creativity else None,
                'res_mot_relationships': float(result.res_mot_relationships) if result.res_mot_relationships else None,
                'res_mot_recognition': float(result.res_mot_recognition) if result.res_mot_recognition else None,
                'res_mot_affiliation': float(result.res_mot_affiliation) if result.res_mot_affiliation else None,
                'res_mot_self_development': float(result.res_mot_self_development) if result.res_mot_self_development else None,
                'res_mot_purpose': float(result.res_mot_purpose) if result.res_mot_purpose else None,
                'res_mot_cooperation': float(result.res_mot_cooperation) if result.res_mot_cooperation else None,
                'res_mot_stability': float(result.res_mot_stability) if result.res_mot_stability else None,
                'res_mot_tradition': float(result.res_mot_tradition) if result.res_mot_tradition else None,
                'res_mot_management': float(result.res_mot_management) if result.res_mot_management else None,
                'res_mot_work_conditions': float(result.res_mot_work_conditions) if result.res_mot_work_conditions else None,
                
                # Ценности (разворачиваем из объекта values)
                'res_val_honesty_justice': result.res_val_honesty_justice,
                'res_val_humanism': result.res_val_humanism,
                'res_val_patriotism': result.res_val_patriotism,
                'res_val_family': result.res_val_family,
                'res_val_health': result.res_val_health,
                'res_val_environment': result.res_val_environment,
            }
            results_list.append(result_data)
            
        return successResponse({
            "student": {
                "stud_id": participant.part_id,
                "stud_name": participant.part_name,
                "stud_gender": participant.part_gender,
                "institution": participant.part_institution.inst_name if participant.part_institution else None,
                "specialty": participant.part_spec.spec_name if participant.part_spec else None,
                "edu_level": participant.part_edu_level.edu_level_name if participant.part_edu_level else None,
                "study_form": participant.part_form.form_name if participant.part_form else None,
                "course_num": participant.part_course_num
            },
            "results": results_list
        })
        
    except ValueError:
        return errorResponse("stud_id must be an integer")
    except Exception as e:
        return exceptionResponse(e)

@csrf_exempt
def value_added_improved(request):
    """
    Улучшенная модель Value-Added для анализа развития компетенций.
    
    Поддерживает два подхода:
    1. Cross-sectional VAM - для всех студентов (сравнение с нормой курса)
    2. Longitudinal VAM - для студентов с повторными замерами (личный прогресс)
    """
    
    session_id = request.GET.get("session_id")
    institution_id = request.GET.get("institution")
    direction_name = request.GET.get("direction")
    course_filter = request.GET.get("course")
    analysis_type = request.GET.get("type", "cross_sectional")  # или "longitudinal"

    # Компетенции
    competencies = [
        "res_comp_leadership",
        "res_comp_communication", 
        "res_comp_self_development",
        "res_comp_result_orientation",
        "res_comp_stress_resistance",
        "res_comp_client_focus",
        "res_comp_planning",
        "res_comp_info_analysis",
        "res_comp_partnership",
        "res_comp_rules_compliance",
        "res_comp_emotional_intel",
        "res_comp_passive_vocab"
    ]

    # Загрузка данных
    results = Results.objects.select_related(
        "res_participant",
        "res_participant__part_institution",
        "res_participant__part_spec"
    ).exclude(
        res_course_num__isnull=True
    ).order_by(
        "res_participant_id",
        "res_year",
        "res_course_num"
    )

    # Фильтрация
    if institution_id:
        results = results.filter(
            res_participant__part_institution__inst_id=institution_id
        )
    if direction_name:
        results = results.filter(
            res_participant__part_spec__spec_name=direction_name
        )
    if course_filter:
        results = results.filter(res_course_num=course_filter)

    results_list = list(results)

    # ========================================
    # CROSS-SECTIONAL VAM (для всех студентов)
    # ========================================
    if analysis_type == "cross_sectional":
        return calculate_cross_sectional_vam(results_list, competencies)
    
    # ========================================
    # LONGITUDINAL VAM (только повторные замеры)
    # ========================================
    elif analysis_type == "longitudinal":
        return calculate_longitudinal_vam(results_list, competencies)
    
    # ========================================
    # COMPARISON (сравнение групп)
    # ========================================
    elif analysis_type == "comparison":
        return calculate_comparison_vam(results_list, competencies)

    return JsonResponse({"status": "error", "message": "Invalid analysis type"})


def calculate_cross_sectional_vam(results_list, competencies):
    """
    Cross-Sectional VAM: Сравнение каждого студента с нормой его курса.
    Работает для ВСЕХ студентов, даже с одним замером.
    """
    
    # 1. Вычисляем средние значения по курсам (эталон)
    course_means = defaultdict(lambda: {comp: [] for comp in competencies})
    
    for row in results_list:
        course = row.res_course_num
        for comp in competencies:
            value = getattr(row, comp)
            if value is not None:
                course_means[course][comp].append(float(value))
    
    # Средние по курсам
    course_baselines = {}
    for course, comp_data in course_means.items():
        course_baselines[course] = {}
        for comp, values in comp_data.items():
            if values:
                course_baselines[course][comp] = np.mean(values)
    
    # 2. Вычисляем VAM для каждого студента
    data = []
    
    for row in results_list:
        course = row.res_course_num
        participant = row.res_participant
        
        if course not in course_baselines:
            continue
        
        vam_by_competency = {}
        vam_values = []
        
        for comp in competencies:
            actual = getattr(row, comp)
            
            if actual is None or comp not in course_baselines[course]:
                continue
            
            expected = course_baselines[course][comp]
            vam = float(actual) - expected
            
            vam_by_competency[comp] = round(vam, 3)
            vam_values.append(vam)
        
        if not vam_values:
            continue
        
        mean_vam = np.mean(vam_values)
        
        data.append({
            "participant_id": row.res_participant_id,
            "participant_name": participant.part_name if participant else "Unknown",
            "year": row.res_year,
            "course": course,
            "mean_vam": round(mean_vam, 3),
            "vam_by_competency": vam_by_competency,
            "institution_id": (
                participant.part_institution.inst_id
                if participant and participant.part_institution else None
            ),
            "institution_name": (
                participant.part_institution.inst_name
                if participant and participant.part_institution else None
            ),
            "direction": (
                participant.part_spec.spec_name
                if participant and participant.part_spec else None
            ),
            "analysis_type": "cross_sectional"
        })
    
    return JsonResponse({
        "status": "success",
        "data": data,
        "total_students": len(data),
        "course_baselines": {
            course: {comp: round(val, 3) for comp, val in baselines.items()}
            for course, baselines in course_baselines.items()
        }
    })


def calculate_longitudinal_vam(results_list, competencies):
    """
    Longitudinal VAM: Отслеживание личного прогресса студентов.
    Работает только для студентов с несколькими замерами.
    """
    
    # Группировка по студентам
    grouped = defaultdict(list)
    for row in results_list:
        grouped[row.res_participant_id].append(row)
    
    # Фильтруем только студентов с 2+ замерами
    grouped = {k: v for k, v in grouped.items() if len(v) >= 2}
    
    # Строим регрессионные модели
    regression_data = {comp: {"x": [], "y": []} for comp in competencies}
    
    for rows in grouped.values():
        for i in range(1, len(rows)):
            prev_row = rows[i - 1]
            curr_row = rows[i]
            
            for comp in competencies:
                prev_score = getattr(prev_row, comp)
                curr_score = getattr(curr_row, comp)
                
                if prev_score is not None and curr_score is not None:
                    regression_data[comp]["x"].append(float(prev_score))
                    regression_data[comp]["y"].append(float(curr_score))
    
    # Вычисляем коэффициенты регрессии
    regression_models = {}
    
    for comp in competencies:
        X = regression_data[comp]["x"]
        Y = regression_data[comp]["y"]
        
        if len(X) < 5:
            continue
        
        n = len(X)
        mean_x = np.mean(X)
        mean_y = np.mean(Y)
        
        numerator = sum((X[i] - mean_x) * (Y[i] - mean_y) for i in range(n))
        denominator = sum((X[i] - mean_x) ** 2 for i in range(n))
        
        if denominator == 0:
            continue
        
        a = numerator / denominator
        b = mean_y - a * mean_x
        
        regression_models[comp] = (a, b)
    
    # Вычисляем VAM для каждого студента
    data = []
    
    for participant_id, rows in grouped.items():
        participant = rows[0].res_participant
        
        for i in range(1, len(rows)):
            prev_row = rows[i - 1]
            curr_row = rows[i]
            
            vam_by_competency = {}
            vam_values = []
            
            for comp in competencies:
                if comp not in regression_models:
                    continue
                
                prev_score = getattr(prev_row, comp)
                curr_score = getattr(curr_row, comp)
                
                if prev_score is None or curr_score is None:
                    continue
                
                a, b = regression_models[comp]
                predicted = a * float(prev_score) + b
                vam = float(curr_score) - predicted
                
                vam_by_competency[comp] = round(vam, 3)
                vam_values.append(vam)
            
            if not vam_values:
                continue
            
            mean_vam = np.mean(vam_values)
            
            data.append({
                "participant_id": participant_id,
                "participant_name": participant.part_name if participant else "Unknown",
                "from_year": prev_row.res_year,
                "to_year": curr_row.res_year,
                "from_course": prev_row.res_course_num,
                "to_course": curr_row.res_course_num,
                "mean_vam": round(mean_vam, 3),
                "vam_by_competency": vam_by_competency,
                "institution_id": (
                    participant.part_institution.inst_id
                    if participant and participant.part_institution else None
                ),
                "institution_name": (
                    participant.part_institution.inst_name
                    if participant and participant.part_institution else None
                ),
                "direction": (
                    participant.part_spec.spec_name
                    if participant and participant.part_spec else None
                ),
                "analysis_type": "longitudinal"
            })
    
    return JsonResponse({
        "status": "success",
        "data": data,
        "total_students": len(grouped),
        "total_transitions": len(data),
        "regression_models": {
            comp: {"a": round(a, 4), "b": round(b, 4)}
            for comp, (a, b) in regression_models.items()
        }
    })


def calculate_comparison_vam(results_list, competencies):
    """
    Сравнительный анализ: Группировка по ВУЗам/направлениям для проверки гипотезы.
    """
    
    # Группировка по институтам и направлениям
    groups = defaultdict(lambda: {comp: [] for comp in competencies})
    
    for row in results_list:
        participant = row.res_participant
        
        if not participant:
            continue
        
        # Ключ группы: (institution, direction, course)
        group_key = (
            participant.part_institution.inst_name if participant.part_institution else "Unknown",
            participant.part_spec.spec_name if participant.part_spec else "Unknown",
            row.res_course_num
        )
        
        for comp in competencies:
            value = getattr(row, comp)
            if value is not None:
                groups[group_key][comp].append(float(value))
    
    # Вычисляем статистику по группам
    data = []
    
    for (institution, direction, course), comp_data in groups.items():
        group_stats = {
            "institution": institution,
            "direction": direction,
            "course": course,
            "student_count": len(comp_data[competencies[0]]) if comp_data[competencies[0]] else 0
        }
        
        # Средние по компетенциям
        competency_means = {}
        all_means = []
        
        for comp in competencies:
            values = comp_data[comp]
            if values:
                mean = np.mean(values)
                competency_means[comp] = round(mean, 3)
                all_means.append(mean)
        
        if all_means:
            group_stats["mean_all_competencies"] = round(np.mean(all_means), 3)
            group_stats["competency_means"] = competency_means
            data.append(group_stats)
    
    return JsonResponse({
        "status": "success",
        "data": data,
        "total_groups": len(data)
    })


# ========================================
# Дополнительная функция для сводной статистики
# ========================================

@csrf_exempt
def vam_summary_statistics(request):
    """
    Сводная статистика для понимания данных и выбора подхода к анализу.
    """
    
    results = Results.objects.select_related(
        "res_participant"
    ).exclude(
        res_course_num__isnull=True
    ).order_by(
        "res_participant_id",
        "res_year"
    )
    
    # Подсчёт замеров по студентам
    student_measurements = defaultdict(int)
    
    for row in results:
        student_measurements[row.res_participant_id] += 1
    
    # Распределение по количеству замеров
    distribution = defaultdict(int)
    for count in student_measurements.values():
        distribution[count] += 1
    
    # Распределение по курсам
    course_distribution = defaultdict(int)
    for row in results:
        course_distribution[row.res_course_num] += 1
    
    # Распределение по годам
    year_distribution = defaultdict(int)
    for row in results:
        year_distribution[row.res_year] += 1
    
    return JsonResponse({
        "status": "success",
        "total_students": len(student_measurements),
        "total_measurements": sum(student_measurements.values()),
        "measurements_distribution": dict(distribution),
        "longitudinal_eligible": sum(1 for c in student_measurements.values() if c >= 2),
        "course_distribution": dict(course_distribution),
        "year_distribution": dict(year_distribution),
        "recommendation": (
            "Рекомендуем использовать Cross-Sectional VAM, так как только "
            f"{sum(1 for c in student_measurements.values() if c >= 2)} студентов "
            f"({round(100 * sum(1 for c in student_measurements.values() if c >= 2) / len(student_measurements), 1)}%) "
            "имеют повторные замеры для Longitudinal VAM."
        )
    })

@method('GET')
@csrf_exempt
def get_filter_options(request):
    """
    Возвращает доступные опции для фильтров (ВУЗы, направления, курсы)
    """
    try:
        session_id = request.GET.get('session_id')
        
        print(f"=== get_filter_options called ===")
        print(f"session_id: {session_id}")
        print(f"Available sessions: {list(data_view_sessions.keys())}")
        
        # Проверка сессии НЕ обязательна для этого endpoint
        # Можем вернуть данные и без сессии
        if session_id and session_id in data_view_sessions:
            session = data_view_sessions[session_id]
            session.update_activity()
            print("Session found and updated")
        else:
            print("No valid session, but continuing anyway")
        
        # Получаем уникальные значения для фильтров
        
        # Институты
        print("Fetching institutions...")
        institutions = Institutions.objects.all().values('inst_id', 'inst_name').order_by('inst_name')
        institutions_list = [
            {'id': inst['inst_id'], 'name': inst['inst_name']} 
            for inst in institutions
        ]
        print(f"Found {len(institutions_list)} institutions")
        if institutions_list:
            print(f"First institution: {institutions_list[0]}")
        
        # Направления (специальности)
        print("Fetching directions...")
        directions = Specialties.objects.all().values_list('spec_name', flat=True).distinct().order_by('spec_name')
        directions_list = list(directions)
        print(f"Found {len(directions_list)} directions")
        if directions_list:
            print(f"First direction: {directions_list[0]}")
        
        # Курсы (обычно 1-6)
        print("Fetching courses...")
        courses = Results.objects.filter(
            res_course_num__isnull=False
        ).values_list('res_course_num', flat=True).distinct().order_by('res_course_num')
        courses_list = list(courses)
        print(f"Found {len(courses_list)} courses: {courses_list}")
        
        # Годы
        print("Fetching years...")
        years = Results.objects.filter(
            res_year__isnull=False
        ).values_list('res_year', flat=True).distinct().order_by('res_year')
        years_list = list(years)
        print(f"Found {len(years_list)} years")
        
        # Центры компетенций
        print("Fetching centers...")
        centers = CompetenceCenters.objects.all().values_list('center_name', flat=True).distinct().order_by('center_name')
        centers_list = list(centers)
        print(f"Found {len(centers_list)} centers")
        
        # Уровни образования
        print("Fetching edu_levels...")
        edu_levels = Educationlevels.objects.all().values_list('edu_level_name', flat=True).distinct().order_by('edu_level_name')
        edu_levels_list = list(edu_levels)
        print(f"Found {len(edu_levels_list)} edu_levels")
        
        # Формы обучения
        print("Fetching study_forms...")
        study_forms = Studyforms.objects.all().values_list('form_name', flat=True).distinct().order_by('form_name')
        study_forms_list = list(study_forms)
        print(f"Found {len(study_forms_list)} study_forms")
        
        response_data = {
            "data": {
                "institutions": institutions_list,
                "directions": directions_list,
                "courses": courses_list,
                "years": years_list,
                "centers": centers_list,
                "edu_levels": edu_levels_list,
                "study_forms": study_forms_list
            }
        }
        
        print(f"Returning response with {len(institutions_list)} institutions, {len(directions_list)} directions, {len(courses_list)} courses")
        
        return successResponse(response_data)
        
    except Exception as e:
        print(f"ERROR in get_filter_options: {str(e)}")
        import traceback
        traceback.print_exc()
        return exceptionResponse(e)


@method('GET')
@csrf_exempt
def get_institution_directions(request):
    """
    Возвращает направления, которые есть в выбранных институтах.
    Извлекает уникальные пары (institution, direction) из таблицы Results.
    """
    try:
        institution_ids = request.GET.getlist('institution_ids[]')
        
        print(f"📊 get_institution_directions called with institutions: {institution_ids}")
        
        # Если институты не выбраны, возвращаем ВСЕ направления
        if not institution_ids or len(institution_ids) == 0:
            directions = Results.objects.filter(
                res_spec__isnull=False
            ).values_list(
                'res_spec__spec_name', flat=True
            ).distinct().order_by('res_spec__spec_name')
            
            directions_list = list(directions)
            print(f"✅ Нет фильтра по ВУЗам, вернули {len(directions_list)} направлений")
            
            return successResponse({
                "directions": directions_list
            })
        
        # Фильтруем направления по выбранным институтам
        directions = Results.objects.filter(
            res_institution__inst_id__in=institution_ids,
            res_spec__isnull=False
        ).values_list(
            'res_spec__spec_name', flat=True
        ).distinct().order_by('res_spec__spec_name')
        
        directions_list = list(directions)
        print(f"✅ Для ВУЗов {institution_ids} нашли {len(directions_list)} направлений")
        
        return successResponse({
            "directions": directions_list
        })
        
    except Exception as e:
        print(f"❌ Ошибка в get_institution_directions: {str(e)}")
        return exceptionResponse(e)


@method('GET')
@csrf_exempt  
def get_vam_comparison(request):
    """
    Получает VAM данные для нескольких институтов/направлений для сравнения.
    Добавлен фильтр по количеству прохождений тестирования.
    """
    try:
        session_id = request.GET.get("session_id")
        analysis_type = request.GET.get("type", "cross_sectional")
        
        # Multi-select параметры
        institution_ids = request.GET.getlist('institution_ids[]')
        directions = request.GET.getlist('directions[]')
        courses = request.GET.getlist('courses[]')
        test_attempts = request.GET.getlist('test_attempts[]')  # Новый параметр!
        
        print(f"\n{'='*60}")
        print(f"📊 get_vam_comparison вызван")
        print(f"   Институты: {institution_ids}")
        print(f"   Направления: {directions}")
        print(f"   Курсы: {courses}")
        print(f"   Прохождений: {test_attempts}")
        print(f"{'='*60}\n")

        competencies = [
            "res_comp_leadership", "res_comp_communication", 
            "res_comp_self_development", "res_comp_result_orientation",
            "res_comp_stress_resistance", "res_comp_client_focus",
            "res_comp_planning", "res_comp_info_analysis",
            "res_comp_partnership", "res_comp_rules_compliance",
            "res_comp_emotional_intel", "res_comp_passive_vocab"
        ]

        # Базовый запрос
        results = Results.objects.select_related(
            "res_participant",
            "res_participant__part_institution",
            "res_participant__part_spec"
        ).exclude(
            res_course_num__isnull=True
        ).order_by(
            "res_participant_id",
            "res_year",
            "res_course_num"
        )

        # Применяем фильтры
        if institution_ids and len(institution_ids) > 0:
            results = results.filter(
                res_participant__part_institution__inst_id__in=institution_ids
            )
        
        if directions and len(directions) > 0:
            results = results.filter(
                res_participant__part_spec__spec_name__in=directions
            )
        
        if courses and len(courses) > 0:
            results = results.filter(res_course_num__in=courses)

        # ============================================================
        # НОВЫЙ ФИЛЬТР: По количеству прохождений тестирования
        # ============================================================
        
        if test_attempts and len(test_attempts) > 0:
            print(f"🔍 Фильтрация по количеству прохождений: {test_attempts}")
            
            # Группируем по студентам и считаем количество замеров
            from django.db.models import Count
            
            # Подсчитываем количество замеров для каждого студента
            student_attempts = Results.objects.values('res_participant').annotate(
                attempt_count=Count('res_id')
            )
            
            # Создаём словарь {student_id: attempt_count}
            attempts_dict = {
                item['res_participant']: item['attempt_count'] 
                for item in student_attempts
            }
            
            # Фильтруем студентов по заданному количеству прохождений
            valid_students = [
                student_id 
                for student_id, count in attempts_dict.items()
                if str(count) in test_attempts
            ]
            
            print(f"✅ Найдено {len(valid_students)} студентов с нужным количеством прохождений")
            
            # Применяем фильтр
            results = results.filter(res_participant__in=valid_students)

        results_list = list(results)
        
        print(f"🔍 Найдено {len(results_list)} записей после всех фильтров")

        # Вычисляем VAM
        if analysis_type == "cross_sectional":
            vam_data = calculate_cross_sectional_vam(results_list, competencies)
        elif analysis_type == "longitudinal":
            vam_data = calculate_longitudinal_vam(results_list, competencies)
        else:
            return JsonResponse({"status": "error", "message": "Invalid analysis type"})
        
        import json
        response_data = json.loads(vam_data.content)
        
        if response_data.get("status") != "success":
            return vam_data
        
        # Группируем данные для графика
        grouped_data = group_vam_for_comparison(
            response_data["data"],
            institution_ids,
            directions
        )
        
        return JsonResponse({
            "status": "success",
            "data": response_data["data"],
            "grouped": grouped_data,
            "total_students": response_data.get("total_students", len(response_data["data"])),
            "filtered_by_attempts": len(test_attempts) > 0,  # Флаг для фронта
            "selected_attempts": test_attempts  # Для отладки
        })
        
    except Exception as e:
        print(f"❌ Ошибка в get_vam_comparison: {str(e)}")
        import traceback
        traceback.print_exc()
        return exceptionResponse(e)


def group_vam_for_comparison(data, institution_ids, directions):
    """
    Группирует VAM данные для построения нескольких линий на графике.
    
    Возвращает:
    {
        "by_institution": {
            "МГУ": {1: [vam1, vam2, ...], 2: [vam3, vam4, ...], ...},
            "СПбГУ": {1: [...], 2: [...], ...}
        },
        "by_direction": {
            "Информатика": {1: [...], 2: [...], ...},
            "Математика": {1: [...], 2: [...], ...}
        }
    }
    """
    
    grouped = {
        "by_institution": defaultdict(lambda: defaultdict(list)),
        "by_direction": defaultdict(lambda: defaultdict(list)),
        "by_institution_direction": defaultdict(lambda: defaultdict(list))
    }
    
    for item in data:
        course = item.get("course") or item.get("to_course")
        vam = item.get("mean_vam", 0)
        institution = item.get("institution_name", "Неизвестно")
        direction = item.get("direction", "Неизвестно")
        
        if course:
            # Группируем по институту
            grouped["by_institution"][institution][course].append(vam)
            
            # Группируем по направлению
            grouped["by_direction"][direction][course].append(vam)
            
            # Группируем по комбинации институт+направление
            key = f"{institution} - {direction}"
            grouped["by_institution_direction"][key][course].append(vam)
    
    # Вычисляем средние значения
    result = {
        "by_institution": {},
        "by_direction": {},
        "by_institution_direction": {}
    }
    
    for inst, courses_data in grouped["by_institution"].items():
        result["by_institution"][inst] = {
            course: round(sum(vams) / len(vams), 2) if vams else 0
            for course, vams in courses_data.items()
        }
    
    for dir, courses_data in grouped["by_direction"].items():
        result["by_direction"][dir] = {
            course: round(sum(vams) / len(vams), 2) if vams else 0
            for course, vams in courses_data.items()
        }
    
    for key, courses_data in grouped["by_institution_direction"].items():
        result["by_institution_direction"][key] = {
            course: round(sum(vams) / len(vams), 2) if vams else 0
            for course, vams in courses_data.items()
        }
    
    return result


@method('GET')
@csrf_exempt
def get_filter_options_with_counts(request):
    """
    Возвращает опции фильтров с количеством записей для каждой.
    С учётом уже выбранных фильтров (cross-filtering).
    
    ВАЖНО: Динамически подсчитывает максимальное количество прохождений!
    """
    try:
        session_id = request.GET.get('session_id')
        

        # Получаем текущие выбранные фильтры
        selected_institution_ids = request.GET.getlist('institution_ids[]')
        selected_directions = request.GET.getlist('directions[]')
        selected_courses = request.GET.getlist('courses[]')
        selected_test_attempts = request.GET.getlist('test_attempts[]')
        
        print(f"\n{'='*60}")
        print(f"📊 get_filter_options_with_counts вызван")
        print(f"   Текущие фильтры:")
        print(f"   - Институты: {selected_institution_ids}")
        print(f"   - Направления: {selected_directions}")
        print(f"   - Курсы: {selected_courses}")
        print(f"   - Прохождения: {selected_test_attempts}")
        print(f"{'='*60}\n")
        
        from django.db.models import Count
        
        # Базовый запрос (применяем все фильтры кроме текущего)
        base_results = Results.objects.select_related(
            'res_participant__part_institution',
            'res_participant__part_spec'
        )
        
        # ============================================================
        # ИНСТИТУТЫ с количеством (с учётом других фильтров)
        # ============================================================
        
        institutions_query = base_results
        
        # Применяем фильтры КРОМЕ институтов
        if selected_directions:
            institutions_query = institutions_query.filter(
                res_participant__part_spec__spec_name__in=selected_directions
            )
        
        if selected_courses:
            institutions_query = institutions_query.filter(
                res_course_num__in=selected_courses
            )
        
        if selected_test_attempts:
            # Фильтр по количеству прохождений
            student_attempts = Results.objects.values('res_participant').annotate(
                attempt_count=Count('res_id')
            )
            
            attempts_dict = {
                item['res_participant']: item['attempt_count'] 
                for item in student_attempts
            }
            
            valid_students = [
                student_id 
                for student_id, count in attempts_dict.items()
                if str(count) in selected_test_attempts
            ]
            
            institutions_query = institutions_query.filter(
                res_participant__in=valid_students
            )
        
        institutions_counts = institutions_query.values(
            'res_participant__part_institution__inst_id',
            'res_participant__part_institution__inst_name'
        ).annotate(
            count=Count('res_id')
        ).order_by('-count')  # Сортировка по количеству (больше → меньше)
        
        institutions_list = [
            {
                'id': item['res_participant__part_institution__inst_id'],
                'name': item['res_participant__part_institution__inst_name'],
                'count': item['count']
            }
            for item in institutions_counts
            if item['res_participant__part_institution__inst_name']
        ]
        
        print(f"✅ Институты: {len(institutions_list)} (отсортировано по количеству)")
        
        # ============================================================
        # НАПРАВЛЕНИЯ с количеством (с учётом других фильтров)
        # ============================================================
        
        directions_query = base_results
        
        # Применяем фильтры КРОМЕ направлений
        if selected_institution_ids:
            directions_query = directions_query.filter(
                res_participant__part_institution__inst_id__in=selected_institution_ids
            )
        
        if selected_courses:
            directions_query = directions_query.filter(
                res_course_num__in=selected_courses
            )
        
        if selected_test_attempts:
            student_attempts = Results.objects.values('res_participant').annotate(
                attempt_count=Count('res_id')
            )
            
            attempts_dict = {
                item['res_participant']: item['attempt_count'] 
                for item in student_attempts
            }
            
            valid_students = [
                student_id 
                for student_id, count in attempts_dict.items()
                if str(count) in selected_test_attempts
            ]
            
            directions_query = directions_query.filter(
                res_participant__in=valid_students
            )
        
        directions_counts = directions_query.values(
            'res_participant__part_spec__spec_name'
        ).annotate(
            count=Count('res_id')
        ).order_by('-count')
        
        directions_list = [
            {
                'id': item['res_participant__part_spec__spec_name'],  # id = name для строк
                'name': item['res_participant__part_spec__spec_name'],
                'count': item['count']
            }
            for item in directions_counts
            if item['res_participant__part_spec__spec_name']
        ]
        
        print(f"✅ Направления: {len(directions_list)} (отсортировано по количеству)")
        
        # ============================================================
        # КУРСЫ с количеством (с учётом других фильтров)
        # ============================================================
        
        courses_query = base_results
        
        # Применяем фильтры КРОМЕ курсов
        if selected_institution_ids:
            courses_query = courses_query.filter(
                res_participant__part_institution__inst_id__in=selected_institution_ids
            )
        
        if selected_directions:
            courses_query = courses_query.filter(
                res_participant__part_spec__spec_name__in=selected_directions
            )
        
        if selected_test_attempts:
            student_attempts = Results.objects.values('res_participant').annotate(
                attempt_count=Count('res_id')
            )
            
            attempts_dict = {
                item['res_participant']: item['attempt_count'] 
                for item in student_attempts
            }
            
            valid_students = [
                student_id 
                for student_id, count in attempts_dict.items()
                if str(count) in selected_test_attempts
            ]
            
            courses_query = courses_query.filter(
                res_participant__in=valid_students
            )
        
        courses_counts = courses_query.values('res_course_num').annotate(
            count=Count('res_id')
        ).order_by('res_course_num')
        
        courses_list = [
            {
                'id': item['res_course_num'],
                'name': f"{item['res_course_num']} курс",
                'count': item['count']
            }
            for item in courses_counts
            if item['res_course_num']
        ]
        
        print(f"✅ Курсы: {len(courses_list)}")
        
        # ============================================================
        # КОЛИЧЕСТВО ПРОХОЖДЕНИЙ (ДИНАМИЧЕСКИЙ ПОДСЧЁТ!)
        # ============================================================
        
        attempts_query = base_results
        
        # Применяем фильтры КРОМЕ прохождений
        if selected_institution_ids:
            attempts_query = attempts_query.filter(
                res_participant__part_institution__inst_id__in=selected_institution_ids
            )
        
        if selected_directions:
            attempts_query = attempts_query.filter(
                res_participant__part_spec__spec_name__in=selected_directions
            )
        
        if selected_courses:
            attempts_query = attempts_query.filter(
                res_course_num__in=selected_courses
            )
        
        # Подсчитываем количество замеров для каждого студента
        student_attempts = attempts_query.values('res_participant').annotate(
            attempt_count=Count('res_id')
        )
        
        # Группируем по количеству прохождений
        from collections import defaultdict
        attempts_distribution = defaultdict(int)
        
        for item in student_attempts:
            attempts_distribution[item['attempt_count']] += 1
        
        # ДИНАМИЧЕСКИ определяем максимум
        max_attempts = max(attempts_distribution.keys()) if attempts_distribution else 6
        
        print(f"🔍 Максимальное количество прохождений: {max_attempts}")
        print(f"   Распределение: {dict(attempts_distribution)}")
        
        # Формируем список (сортировка по количеству прохождений)
        test_attempts_list = [
            {
                'id': attempts,
                'name': f"{attempts} прохождение" if attempts == 1 else f"{attempts} прохождения",
                'count': students_count
            }
            for attempts, students_count in sorted(attempts_distribution.items())
        ]
        
        print(f"✅ Прохождений: {len(test_attempts_list)} (от 1 до {max_attempts})")
        
        # Компетенции с русскими названиями
        competencies_data = [
            {"id": "res_comp_info_analysis", "name": "Анализ информации"},
            {"id": "res_comp_planning", "name": "Планирование"},
            {"id": "res_comp_result_orientation", "name": "Ориентация на результат"},
            {"id": "res_comp_stress_resistance", "name": "Стрессоустойчивость"},
            {"id": "res_comp_partnership", "name": "Партнёрство"},
            {"id": "res_comp_rules_compliance", "name": "Соблюдение правил"},
            {"id": "res_comp_self_development", "name": "Саморазвитие"},
            {"id": "res_comp_leadership", "name": "Лидерство"},
            {"id": "res_comp_emotional_intel", "name": "Эмоциональный интеллект"},
            {"id": "res_comp_client_focus", "name": "Клиентоориентированность"},
            {"id": "res_comp_communication", "name": "Коммуникация"},
            {"id": "res_comp_passive_vocab", "name": "Пассивный словарь"},
        ]

        competencies_list = []

        for comp_info in competencies_data:
            comp_field = comp_info["id"]
            
            # Подсчитываем ненулевые значения
            count = Results.objects.exclude(
                Q(**{f"{comp_field}__isnull": True}) | Q(**{comp_field: 0})
            ).count()
            
            competencies_list.append({
                "id": comp_field,
                "name": comp_info["name"],
                "count": count
            })

        print(f"✅ Компетенции: {len(competencies_list)}")


        students_query = Participants.objects.annotate(
            results_count=Count('results')
        ).filter(
            results_count__gt=0
        ).order_by('part_name')[:1000]  # Ограничиваем 1000 для производительности

        students_list = [
            {
                'id': student.part_id,
                'name': f"{student.part_name} (ID: {student.part_id})",
                'count': student.results_count
            }
            for student in students_query
        ]

        print(f"✅ Студенты: {len(students_list)}")

        return successResponse({
            "data": {
                "institutions": institutions_list,
                "directions": directions_list,
                "courses": courses_list,
                "test_attempts": test_attempts_list,
                "competencies": competencies_list,  # ← С count!
                "students": students_list,  # ← ДОБАВЛЕНО!
                "max_attempts": max_attempts
            }
        })
        
    except Exception as e:
        print(f"❌ Ошибка в get_filter_options_with_counts: {str(e)}")
        import traceback
        traceback.print_exc()
        return exceptionResponse(e)


# ============================================================
# UNIFIED VAM FUNCTION (вместо Cross-Sectional и Longitudinal)
# ============================================================

@method('GET')
@csrf_exempt  
def get_vam_unified(request):
    """
    ЕДИНЫЙ метод VAM, который автоматически определяет тип анализа
    на основе фильтра "Количество прохождений":
    
    - Если выбрано "1 прохождение" → Cross-Sectional VAM
    - Если выбрано "2+" прохождений → Longitudinal VAM
    - Если не выбрано → Mixed (оба подхода)
    """
    try:
        session_id = request.GET.get("session_id")
        
        # Фильтры
        institution_ids = request.GET.getlist('institution_ids[]')
        directions = request.GET.getlist('directions[]')
        courses = request.GET.getlist('courses[]')
        test_attempts = request.GET.getlist('test_attempts[]')
        selected_competencies = request.GET.getlist('competencies[]')
        student_ids = request.GET.getlist('student_ids[]')
        
        print(f"\n{'='*60}")
        print(f"📊 get_vam_unified вызван")
        print(f"   Институты: {institution_ids}")
        print(f"   Направления: {directions}")
        print(f"   Курсы: {courses}")
        print(f"   Прохождений: {test_attempts}")
        print(f"   Студенты: {student_ids}")
        print(f"{'='*60}\n")

        competencies = [
            "res_comp_leadership", "res_comp_communication", 
            "res_comp_self_development", "res_comp_result_orientation",
            "res_comp_stress_resistance", "res_comp_client_focus",
            "res_comp_planning", "res_comp_info_analysis",
            "res_comp_partnership", "res_comp_rules_compliance",
            "res_comp_emotional_intel", "res_comp_passive_vocab"
        ]

        # Базовый запрос
        results = Results.objects.select_related(
            "res_participant",
            "res_participant__part_institution",
            "res_participant__part_spec"
        ).exclude(
            res_course_num__isnull=True
        ).order_by(
            "res_participant_id",
            "res_year",
            "res_course_num"
        )

        # Применяем фильтры
        if institution_ids and len(institution_ids) > 0:
            results = results.filter(
                res_participant__part_institution__inst_id__in=institution_ids
            )
        
        if directions and len(directions) > 0:
            results = results.filter(
                res_participant__part_spec__spec_name__in=directions
            )
        
        if courses and len(courses) > 0:
            results = results.filter(res_course_num__in=courses)

        # Фильтр по количеству прохождений
        if test_attempts and len(test_attempts) > 0:
            print(f"🔍 Фильтрация по количеству прохождений: {test_attempts}")
            
            from django.db.models import Count
            
            student_attempts = Results.objects.values('res_participant').annotate(
                attempt_count=Count('res_id')
            )
            
            attempts_dict = {
                item['res_participant']: item['attempt_count'] 
                for item in student_attempts
            }
            
            valid_students = [
                student_id 
                for student_id, count in attempts_dict.items()
                if str(count) in test_attempts
            ]
            
            print(f"✅ Найдено {len(valid_students)} студентов")
            
            results = results.filter(res_participant__in=valid_students)

        if selected_competencies:
            # Фильтруем список компетенций
            competencies = [c for c in competencies if c in selected_competencies]

        if student_ids and len(student_ids) > 0:
            # Конвертируем в int, так как part_id это число
            student_ids_int = [int(sid) for sid in student_ids if sid.isdigit()]
            results = results.filter(res_participant__part_id__in=student_ids_int)
            print(f"   → Фильтр: {len(student_ids_int)} студентов")

        results_list = list(results)
        
        print(f"🔍 Найдено {len(results_list)} записей после всех фильтров")

        # ============================================================
        # АВТОМАТИЧЕСКОЕ ОПРЕДЕЛЕНИЕ ТИПА АНАЛИЗА
        # ============================================================
        
        # Если выбрано только "1 прохождение" → Cross-Sectional
        if test_attempts == ['1']:
            print("📊 Используем Cross-Sectional VAM (только 1 прохождение)")
            vam_data = calculate_cross_sectional_vam(results_list, competencies)
            analysis_method = "cross_sectional"
        
        # Если выбраны только "2+" прохождения → Longitudinal
        elif test_attempts and all(int(a) >= 2 for a in test_attempts):
            print("📊 Используем Longitudinal VAM (2+ прохождения)")
            vam_data = calculate_longitudinal_vam(results_list, competencies)
            analysis_method = "longitudinal"
        
        # Если смешанные или не выбрано → Mixed (приоритет Longitudinal)
        else:
            print("📊 Используем Mixed VAM (приоритет Longitudinal)")
            vam_data = calculate_longitudinal_vam(results_list, competencies)
            analysis_method = "mixed"
        
        import json
        response_data = json.loads(vam_data.content)
        
        if response_data.get("status") != "success":
            return vam_data
        
        # Группируем данные для графика
        grouped_data = group_vam_for_comparison(
            response_data["data"],
            institution_ids,
            directions
        )
        
        return JsonResponse({
            "status": "success",
            "data": response_data["data"],
            "grouped": grouped_data,
            "total_students": response_data.get("total_students", len(response_data["data"])),
            "analysis_method": analysis_method,  # Для информации на фронте
            "selected_attempts": test_attempts,
            "selected_competencies": selected_competencies
        })
        
    except Exception as e:
        print(f"❌ Ошибка в get_vam_unified: {str(e)}")
        import traceback
        traceback.print_exc()
        return exceptionResponse(e)

@method('GET')
@csrf_exempt
def get_latent_growth(request):
    """
    Получает данные Latent Growth Model (LGM) с фильтрацией.
    ОБНОВЛЕНО: Теперь поддерживает группировку по институтам/направлениям!
    """
    try:
        # Получаем фильтры
        institution_ids = request.GET.getlist('institution_ids[]')
        directions = request.GET.getlist('directions[]')
        courses = request.GET.getlist('courses[]')
        test_attempts = request.GET.getlist('test_attempts[]')
        selected_competencies = request.GET.getlist('competencies[]')

        print(f"\n{'='*60}")
        print(f"📊 get_latent_growth вызван")
        print(f"   Институты: {institution_ids}")
        print(f"   Направления: {directions}")
        print(f"{'='*60}\n")

        # Список всех компетенций
        all_competencies = [
            "res_comp_leadership", "res_comp_communication",
            "res_comp_self_development", "res_comp_result_orientation",
            "res_comp_stress_resistance", "res_comp_client_focus",
            "res_comp_planning", "res_comp_info_analysis",
            "res_comp_partnership", "res_comp_rules_compliance",
            "res_comp_emotional_intel", "res_comp_passive_vocab"
        ]

        # Фильтруем компетенции
        if selected_competencies:
            competencies = [c for c in all_competencies if c in selected_competencies]
        else:
            competencies = all_competencies

        # Базовый запрос
        results = Results.objects.select_related(
            "res_participant",
            "res_participant__part_institution",
            "res_participant__part_spec"
        ).exclude(res_course_num__isnull=True)

        # Применяем фильтры
        if institution_ids:
            results = results.filter(
                res_participant__part_institution__inst_id__in=institution_ids
            )

        if directions:
            results = results.filter(
                res_participant__part_spec__spec_name__in=directions
            )

        if courses:
            results = results.filter(res_course_num__in=courses)

        # Фильтр по количеству прохождений
        if test_attempts:
            from django.db.models import Count
            
            student_attempts = Results.objects.values('res_participant').annotate(
                attempt_count=Count('res_id')
            )
            
            attempts_dict = {
                item['res_participant']: item['attempt_count'] 
                for item in student_attempts
            }
            
            valid_students = [
                student_id 
                for student_id, count in attempts_dict.items()
                if str(count) in test_attempts
            ]
            
            results = results.filter(res_participant__in=valid_students)

        results_list = list(results)
        
        print(f"🔍 Итого записей: {len(results_list)}")

        # ============================================================
        # ОПРЕДЕЛЯЕМ ТИП ГРУППИРОВКИ (как в VAM!)
        # ============================================================
        
        group_by = None
        
        if len(institution_ids) > 0 and len(directions) == 0:
            group_by = 'by_institution'
            print(f"   Группировка: по институтам ({len(institution_ids)})")
        elif len(directions) > 0 and len(institution_ids) == 0:
            group_by = 'by_direction'
            print(f"   Группировка: по направлениям ({len(directions)})")
        elif len(institution_ids) > 0 and len(directions) > 0:
            group_by = 'by_institution_direction'
            print(f"   Группировка: институт + направление")
        else:
            group_by = 'overall'
            print(f"   Группировка: общая")

        # ============================================================
        # ВЫЧИСЛЯЕМ ТРАЕКТОРИИ С ГРУППИРОВКОЙ
        # ============================================================
        
        if group_by == 'overall':
            # Без группировки - как раньше
            growth_data = calculate_population_growth(results_list, competencies)
        else:
            # С группировкой - новая функция!
            growth_data = calculate_grouped_growth(
                results_list, 
                competencies,
                group_by
            )

        # Добавляем метаданные
        if growth_data.get("status") == "success":
            growth_data["group_by"] = group_by
            growth_data["filters_applied"] = {
                "institutions": len(institution_ids),
                "directions": len(directions),
                "courses": len(courses),
                "test_attempts": len(test_attempts),
                "competencies": len(competencies)
            }

        return JsonResponse(growth_data)

    except Exception as e:
        print(f"❌ Ошибка в get_latent_growth: {str(e)}")
        import traceback
        traceback.print_exc()
        return exceptionResponse(e)

def calculate_population_growth(results_list, competencies):
    """
    Latent Growth Model (LGM) - модель скрытого роста.
    Вычисляет средние траектории развития компетенций на уровне популяции.
    
    Возвращает среднее значение каждой компетенции по курсам.
    """
    try:
        from collections import defaultdict
        
        print(f"\n{'='*60}")
        print(f"📊 calculate_population_growth вызван")
        print(f"   Записей: {len(results_list)}")
        print(f"   Компетенций: {len(competencies)}")
        print(f"{'='*60}\n")
        
        # Группируем по курсам
        by_course = defaultdict(lambda: defaultdict(list))
        
        for result in results_list:
            course = result.res_course_num
            if not course:
                continue
            
            for comp in competencies:
                value = getattr(result, comp, None)
                if value is not None:
                    by_course[course][comp].append(value)
        
        # Вычисляем средние для каждого курса
        growth_trajectory = {}
        
        for comp in competencies:
            trajectory = []
            
            for course in sorted(by_course.keys()):
                if comp in by_course[course] and len(by_course[course][comp]) > 0:
                    values = by_course[course][comp]
                    avg = sum(values) / len(values)
                    trajectory.append({
                        "course": course,
                        "mean": round(avg, 2),
                        "count": len(values)
                    })
            
            if trajectory:  # Только если есть данные
                growth_trajectory[comp] = trajectory
                print(f"   ✅ {comp}: {len(trajectory)} точек данных")
        
        print(f"\n✅ Построены траектории для {len(growth_trajectory)} компетенций")
        
        return {
            "status": "success",
            "data": growth_trajectory,
            "model": "latent_growth",
            "total_records": len(results_list),
            "competencies_count": len(growth_trajectory)
        }
    
    except Exception as e:
        print(f"❌ Ошибка в calculate_population_growth: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            "status": "error",
            "message": str(e)
        }
    
def calculate_grouped_growth(results_list, competencies, group_by):
    """
    Вычисляет траектории роста с группировкой по институтам/направлениям.
    Возвращает данные в формате для multi-line графиков.
    """
    try:
        from collections import defaultdict
        
        print(f"\n📊 calculate_grouped_growth: {group_by}")
        
        # Группируем результаты
        groups = defaultdict(list)
        
        for result in results_list:
            # Определяем группу
            if group_by == 'by_institution':
                group_name = result.res_participant.part_institution.inst_name if result.res_participant.part_institution else "Неизвестно"
            elif group_by == 'by_direction':
                group_name = result.res_participant.part_spec.spec_name if result.res_participant.part_spec else "Неизвестно"
            elif group_by == 'by_institution_direction':
                inst = result.res_participant.part_institution.inst_name if result.res_participant.part_institution else "Неизвестно"
                spec = result.res_participant.part_spec.spec_name if result.res_participant.part_spec else "Неизвестно"
                group_name = f"{inst} - {spec}"
            else:
                group_name = "Все"
            
            groups[group_name].append(result)
        
        print(f"   Найдено групп: {len(groups)}")
        for group_name, group_results in list(groups.items())[:3]:
            print(f"   - {group_name}: {len(group_results)} записей")
        
        # Вычисляем траектории для каждой компетенции
        growth_trajectory = {}
        
        for comp in competencies:
            comp_data = {}
            
            # Для каждой группы вычисляем траекторию
            for group_name, group_results in groups.items():
                by_course = defaultdict(list)
                
                for result in group_results:
                    course = result.res_course_num
                    if not course:
                        continue
                    
                    value = getattr(result, comp, None)
                    if value is not None:
                        by_course[course].append(value)
                
                # Вычисляем средние по курсам
                trajectory = []
                for course in sorted(by_course.keys()):
                    if by_course[course]:
                        values = by_course[course]
                        avg = sum(values) / len(values)
                        trajectory.append({
                            "course": course,
                            "mean": round(avg, 2),
                            "count": len(values)
                        })
                
                if trajectory:
                    comp_data[group_name] = trajectory
            
            if comp_data:
                growth_trajectory[comp] = comp_data
        
        print(f"✅ Построены траектории для {len(growth_trajectory)} компетенций")
        
        return {
            "status": "success",
            "data": growth_trajectory,
            "model": "latent_growth",
            "group_by": group_by,
            "total_records": len(results_list),
            "groups_count": len(groups)
        }
    
    except Exception as e:
        print(f"❌ Ошибка в calculate_grouped_growth: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            "status": "error",
            "message": str(e)
        }